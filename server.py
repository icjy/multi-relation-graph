from __future__ import annotations

import base64
import csv
import io
import json
import math
import mimetypes
import os
import re
import subprocess
import sys
import tempfile
import threading
import traceback
import uuid
from collections import defaultdict
from datetime import date, datetime, time
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from itertools import combinations
from pathlib import Path
from decimal import Decimal
from typing import Any

os.environ.setdefault("HF_ENDPOINT", "https://hf-mirror.com")
os.environ.setdefault("HF_HUB_OFFLINE", "1")

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - handled at runtime
    load_workbook = None

try:
    import igraph as ig
except ImportError:  # pragma: no cover - handled at runtime
    ig = None

try:
    import leidenalg
except ImportError:  # pragma: no cover - handled at runtime
    leidenalg = None

try:
    import numpy as np
except ImportError:  # pragma: no cover - handled at runtime
    np = None

try:
    from sentence_transformers import SentenceTransformer
except ImportError:  # pragma: no cover - handled at runtime
    SentenceTransformer = None

try:
    from sklearn.cluster import DBSCAN
except ImportError:  # pragma: no cover - handled at runtime
    DBSCAN = None

try:
    from sklearn.preprocessing import normalize as sklearn_normalize
except ImportError:  # pragma: no cover - handled at runtime
    sklearn_normalize = None

try:
    import faiss
except ImportError:  # pragma: no cover - handled at runtime
    faiss = None

try:
    import pandas as pd
except ImportError:  # pragma: no cover - handled at runtime
    pd = None

try:
    import cpca
except ImportError:  # pragma: no cover - handled at runtime
    cpca = None

try:
    import cn2an
except ImportError:  # pragma: no cover - handled at runtime
    cn2an = None


ROOT = Path(__file__).resolve().parent
STATIC = ROOT / "static"
# HOST = "127.0.0.1"
HOST = "0.0.0.0"
PORT = int(os.environ.get("PORT", "8000"))

REQUIRED_COLUMNS = {
    "loan_task_id",
    "apply_time",
    "final_result",
    "reloan_flag",
    "return_flag",
    "app_key",
    "app_user_id",
    "consigneeMobileId",
    "ip",
    "device_id",
    "receiverAddr",
    "addr_cluster_id",
    "fpd1",
    "fpd3",
    "fpd10",
    "fpd30",
    "cpd1",
    "cpd7",
    "cpd10",
    "cpd30",
    "spd10",
    "spd30",
    "tpd10",
    "tpd30",
}
OVERDUE_FIELDS = (
    "fpd1",
    "fpd3",
    "fpd7",
    "fpd10",
    "fpd30",
    "cpd1",
    "cpd7",
    "cpd10",
    "cpd30",
    "spd10",
    "spd30",
    "tpd10",
    "tpd30",
)
AGENT_COL = "收货人手机号"
BORROWER_COL = "借款人手机号"
ADDRESS_COL = "收货人地址"
ADDR_CLUSTER_COL = "addr_clusterid"
DEVICE_COL = "设备号"
IP_COL = "ip"
RELOAN_COL = "reloan_flag"
RETURN_COL = "return_flag"
FINAL_RESULT_COL = "final_result"
FUNDED_RESULT = "30"
QUERY_GRAPH_MAX_NODES = 80

FIELD_ALIASES = {
    AGENT_COL: (AGENT_COL, "consigneeMobileId"),
    BORROWER_COL: (BORROWER_COL, "app_user_id"),
    ADDRESS_COL: (ADDRESS_COL, "receiverAddr", "consigneeAddr"),
    ADDR_CLUSTER_COL: (ADDR_CLUSTER_COL, "addr_cluster_id"),
    DEVICE_COL: (DEVICE_COL, "device_id"),
    IP_COL: (IP_COL,),
    RELOAN_COL: (RELOAN_COL,),
    RETURN_COL: (RETURN_COL,),
    FINAL_RESULT_COL: (FINAL_RESULT_COL,),
}

DATASETS: dict[str, dict[str, Any]] = {}
CLUSTER_JOBS: dict[str, dict[str, Any]] = {}
CLUSTER_JOBS_LOCK = threading.Lock()
EMBEDDING_MODEL_NAME = "paraphrase-multilingual-MiniLM-L12-v2"
EMBEDDING_MODEL: Any = None
EMBEDDING_MODEL_LOCK = threading.Lock()
EMBEDDING_CACHE: dict[str, list[float]] = {}


def json_response(handler: BaseHTTPRequestHandler, status: int, payload: dict[str, Any]) -> None:
    body = json.dumps(payload, ensure_ascii=False, default=json_default).encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "application/json; charset=utf-8")
    handler.send_header("Content-Length", str(len(body)))
    handler.end_headers()
    handler.wfile.write(body)


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().replace("\ufeff", "")


def json_safe_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, float) and math.isnan(value):
        return ""
    return value


def json_default(value: Any) -> Any:
    converted = json_safe_value(value)
    if converted is not value:
        return converted
    return str(value)


def normalize_phone(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if math.isnan(value):
            return ""
        if value.is_integer():
            value = int(value)
    text = str(value).strip()
    if not text or text.lower() in {"nan", "none", "null"}:
        return ""
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    digits = re.sub(r"\D", "", text)
    return digits or text


def normalize_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if math.isnan(value):
            return ""
        if value.is_integer():
            value = int(value)
    text = str(value).strip()
    if not text or text.lower() in {"nan", "none", "null"}:
        return ""
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    return text


def clean_receiver_address(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    text = str(value).lower()
    text = re.sub(r"[\-]", "", text)
    text = re.sub(r"\s+", "", text)
    text = text.replace("\u3000", "")
    text = text.replace("（", "(").replace("）", ")")
    text = text.replace("：", ":").replace(":", "")
    text = text.replace("(", "").replace(")", "")
    text = re.sub(r"(地址|电话|姓名)", "", text)

    if cpca is None or pd is None:
        return text

    try:
        parsed = cpca.transform([text])
        row = parsed.iloc[0]
        province = row["省"]
        city = row["市"]
        district = row["区"]
        detail = row["地址"]
        admin_words = get_admin_words(province, city, district, None)
        detail = remove_admin_repeat(detail, admin_words)
        detail = normalize_address_number(detail)
        return "".join(
            [
                "" if pd.isna(province) else str(province),
                "" if pd.isna(city) else str(city),
                "" if pd.isna(district) else str(district),
                detail,
            ]
        )
    except Exception:
        return text


def get_admin_words(province: Any, city: Any, district: Any, town: Any) -> list[str]:
    if pd is None:
        return []
    words: list[str] = []
    for value in (province, city, district, town):
        if pd.isna(value):
            continue
        text = str(value)
        words.append(text)
        short = (
            text.replace("省", "")
            .replace("市", "")
            .replace("区", "")
            .replace("县", "")
            .replace("自治州", "")
            .replace("街道", "")
            .replace("镇", "")
        )
        words.append(short)
    return sorted(set(words), key=len, reverse=True)


def remove_admin_repeat(detail: Any, admin_words: list[str]) -> str:
    text = "" if detail is None else str(detail)
    changed = True
    while changed:
        changed = False
        for word in admin_words:
            if word and text.startswith(word):
                text = text[len(word) :]
                changed = True
                break
    return text


def normalize_address_number(text: Any) -> str:
    value = "" if text is None else str(text)
    if cn2an is None:
        return value
    try:
        return cn2an.transform(value, "cn2an")
    except Exception:
        return value


def require_clustering_dependencies(incremental: bool = False) -> None:
    missing: list[str] = []
    if np is None:
        missing.append("numpy")
    if SentenceTransformer is None:
        missing.append("sentence-transformers")
    if DBSCAN is None:
        missing.append("scikit-learn")
    if sklearn_normalize is None:
        missing.append("scikit-learn")
    if pd is None:
        missing.append("pandas")
    if cpca is None:
        missing.append("cpca")
    if cn2an is None:
        missing.append("cn2an")
    if incremental and faiss is None:
        missing.append("faiss-cpu")
    if missing:
        raise ValueError(f"当前环境缺少地址聚类依赖：{', '.join(sorted(set(missing)))}。请执行：python3 -m pip install -r requirements.txt")


def short_error(exc: Exception) -> str:
    text = str(exc).replace("\n", " ").strip()
    return text[:700] + ("..." if len(text) > 700 else "")


def short_text(value: Any, limit: int = 700) -> str:
    text = str(value).replace("\n", " ").strip()
    return text[:limit] + ("..." if len(text) > limit else "")


def embedding_model() -> Any:
    global EMBEDDING_MODEL
    if SentenceTransformer is None:
        raise ValueError("当前环境缺少 sentence-transformers。请执行：python3 -m pip install -r requirements.txt")
    with EMBEDDING_MODEL_LOCK:
        if EMBEDDING_MODEL is None:
            try:
                EMBEDDING_MODEL = SentenceTransformer(EMBEDDING_MODEL_NAME, local_files_only=True)
            except Exception as exc:  # noqa: BLE001 - model download/load errors need UI-friendly text
                raise ValueError(
                    "地址 embedding 模型加载失败，增量新地址无法编码。"
                    f"模型：{EMBEDDING_MODEL_NAME}；当前使用本地离线加载，未联网下载；"
                    f"原始原因：{short_error(exc)}"
                ) from exc
        return EMBEDDING_MODEL


def embed_addresses(addresses: list[str], job_id: str | None = None) -> Any:
    require_clustering_dependencies()
    if np is None:
        raise ValueError("当前环境缺少 numpy。")
    model = embedding_model()
    vectors: list[list[float] | None] = []
    missing_addresses: list[str] = []
    missing_positions: list[int] = []
    for index, address in enumerate(addresses):
        cached = EMBEDDING_CACHE.get(address)
        if cached is None:
            vectors.append(None)
            missing_addresses.append(address)
            missing_positions.append(index)
        else:
            vectors.append(cached)
    if missing_addresses:
        update_cluster_job(job_id, progress=18, message=f"正在编码地址向量：0/{len(missing_addresses)}")
        encoded = model.encode(missing_addresses, show_progress_bar=False)
        for offset, vector in enumerate(encoded):
            values = np.asarray(vector, dtype="float32").tolist()
            EMBEDDING_CACHE[missing_addresses[offset]] = values
            vectors[missing_positions[offset]] = values
            if job_id and (offset + 1 == len(missing_addresses) or (offset + 1) % 200 == 0):
                progress = 18 + int((offset + 1) / len(missing_addresses) * 42)
                update_cluster_job(job_id, progress=progress, message=f"正在编码地址向量：{offset + 1}/{len(missing_addresses)}")
    return np.asarray(vectors, dtype="float32")


def rows_with_receiver_addr(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for row in rows:
        aliased = apply_aliases(row)
        clean_row = {key: json_safe_value(value) for key, value in row.items()}
        receiver = clean_receiver_address(get_field(aliased, ADDRESS_COL))
        clean_row["receiverAddr"] = receiver
        result.append(clean_row)
    return result


def build_cluster_tables(rows: list[dict[str, Any]], labels: list[int]) -> dict[str, Any]:
    detail_rows: list[dict[str, Any]] = []
    summary: dict[str, dict[str, Any]] = {}
    source_columns: list[str] = list(rows[0].keys()) if rows else []
    columns = source_columns[:]
    for column in ("receiverAddr", "addr_cluster_id"):
        if column not in columns:
            columns.append(column)
    for row, label in zip(rows, labels):
        detail = {key: json_safe_value(value) for key, value in row.items()}
        receiver = clean_receiver_address(detail.get("receiverAddr") or detail.get("consigneeAddr") or get_field(apply_aliases(row), ADDRESS_COL))
        cluster_id = int(label)
        detail["receiverAddr"] = receiver
        detail["addr_cluster_id"] = cluster_id
        detail_rows.append(detail)
        key = str(cluster_id)
        if key not in summary:
            summary[key] = {"addr_cluster_id": cluster_id, "sample_receiverAddr": receiver, "count": 0}
        summary[key]["count"] += 1
        if not summary[key]["sample_receiverAddr"] and receiver:
            summary[key]["sample_receiverAddr"] = receiver
    summary_rows = sorted(summary.values(), key=lambda item: (-item["count"], item["addr_cluster_id"]))
    return {
        "detail_table": {"columns": columns, "rows": detail_rows},
        "summary_table": {"columns": ["addr_cluster_id", "sample_receiverAddr", "count"], "rows": summary_rows},
    }


def engineering_payload(rows: list[dict[str, Any]], labels: list[int], embeddings: Any, params: dict[str, Any]) -> dict[str, Any]:
    vectors = embeddings.tolist() if hasattr(embeddings, "tolist") else embeddings
    items = []
    for row, label, vector in zip(rows, labels, vectors):
        receiver = clean_receiver_address(row.get("receiverAddr") or row.get("consigneeAddr") or get_field(apply_aliases(row), ADDRESS_COL))
        items.append(
            {
                "receiverAddr": receiver,
                "addr_cluster_id": int(label),
                "embedding": vector,
            }
        )
    return {
        "version": 1,
        "model": EMBEDDING_MODEL_NAME,
        "params": params,
        "items": items,
    }


def get_field(row: dict[str, Any], canonical: str) -> Any:
    for column in FIELD_ALIASES.get(canonical, (canonical,)):
        if column in row and row.get(column) not in (None, ""):
            return row.get(column)
    return ""


def apply_aliases(row: dict[str, Any]) -> dict[str, Any]:
    result = {key: json_safe_value(value) for key, value in row.items()}
    for canonical in FIELD_ALIASES:
        value = get_field(result, canonical)
        if value not in (None, ""):
            result[canonical] = json_safe_value(value)
    return result


def filter_rows(rows: list[dict[str, Any]], reloan_filter: str = "all", return_filter: str = "all") -> list[dict[str, Any]]:
    reloan_filter = normalize_value(reloan_filter or "all")
    return_filter = normalize_value(return_filter or "all")
    filtered: list[dict[str, Any]] = []
    for row in rows:
        if reloan_filter != "all" and normalize_value(get_field(row, RELOAN_COL)) != reloan_filter:
            continue
        if return_filter != "all" and normalize_value(get_field(row, RETURN_COL)) != return_filter:
            continue
        filtered.append(row)
    return filtered


def filter_values(rows: list[dict[str, Any]], canonical: str) -> list[str]:
    values = {
        normalize_value(get_field(apply_aliases(row), canonical))
        for row in rows
        if normalize_value(get_field(apply_aliases(row), canonical))
    }
    return sorted(values, key=lambda value: (value.lower(), value))


def build_filter_options(rows: list[dict[str, Any]]) -> dict[str, list[str]]:
    return {
        "reloan": filter_values(rows, RELOAN_COL),
        "return": filter_values(rows, RETURN_COL),
        "final_result": filter_values(rows, FINAL_RESULT_COL),
    }


def complex_query_rows(
    rows: list[dict[str, Any]],
    filters: dict[str, Any] | None = None,
    query_values: dict[str, Any] | None = None,
) -> dict[str, Any]:
    filters = filters or {}
    query_values = query_values or {}
    query_fields = {
        "app_user_id": BORROWER_COL,
        "consigneeMobileId": AGENT_COL,
        "device_id": DEVICE_COL,
        "ip": IP_COL,
        "addr_cluster_id": ADDR_CLUSTER_COL,
    }
    parsed_values = {
        field: {
            normalize_value(item)
            for item in str(raw_value or "").split("|")
            if normalize_value(item)
        }
        for field, raw_value in query_values.items()
        if field in query_fields
    }
    parsed_values = {field: values for field, values in parsed_values.items() if values}
    input_counts = {field: len(values) for field, values in parsed_values.items()}
    matched_values = {field: set() for field in parsed_values}

    def filter_match(row: dict[str, Any]) -> bool:
        checks = (
            (RELOAN_COL, filters.get("reloan_flag")),
            (RETURN_COL, filters.get("return_flag")),
            (FINAL_RESULT_COL, filters.get("final_result")),
        )
        for canonical, expected in checks:
            expected = normalize_value(expected or "all")
            if expected != "all" and normalize_value(get_field(row, canonical)) != expected:
                return False
        return True

    def query_match(row: dict[str, Any]) -> bool:
        if not parsed_values:
            return True
        matched = False
        for field, values in parsed_values.items():
            canonical = query_fields[field]
            value = normalize_value(get_field(row, canonical))
            if value and value in values:
                matched_values[field].add(value)
                matched = True
        return matched

    filtered_rows = [row for row in rows if filter_match(row)]
    matched_rows = [row for row in filtered_rows if query_match(row)]
    aliased_rows = [apply_aliases(row) for row in matched_rows]

    def distinct_count(canonical: str) -> int:
        return len({
            normalize_value(get_field(row, canonical))
            for row in aliased_rows
            if normalize_value(get_field(row, canonical))
        })

    def distinct_raw(column: str) -> int:
        return len({
            normalize_value(row.get(column))
            for row in matched_rows
            if normalize_value(row.get(column))
        })

    unmatched_values = {
        field: sorted(values - matched_values.get(field, set()), key=lambda item: (item.lower(), item))
        for field, values in parsed_values.items()
    }
    matched_value_counts = {field: len(matched_values.get(field, set())) for field in parsed_values}

    summary = {
        "row_count": len(matched_rows),
        "loan_task_id_count": distinct_raw("loan_task_id"),
        "borrower_count": distinct_count(BORROWER_COL),
        "agent_count": distinct_count(AGENT_COL),
        "device_count": distinct_count(DEVICE_COL),
        "ip_count": distinct_count(IP_COL),
        "addr_cluster_count": distinct_count(ADDR_CLUSTER_COL),
        "raw_address_count": distinct_raw("consigneeAddr"),
        "clean_address_count": distinct_raw("receiverAddr"),
    }
    columns = list(rows[0].keys()) if rows else []
    return {
        "columns": columns,
        "rows": [{key: json_safe_value(value) for key, value in row.items()} for row in matched_rows],
        "summary": summary,
        "input_counts": input_counts,
        "matched_value_counts": matched_value_counts,
        "unmatched_values": unmatched_values,
        "relation_graph": build_query_relation_graph(filtered_rows, parsed_values),
    }


def build_query_relation_graph(rows: list[dict[str, Any]], parsed_values: dict[str, set[str]]) -> dict[str, Any]:
    if parsed_values.get("consigneeMobileId"):
        return build_agent_query_graph(rows, parsed_values["consigneeMobileId"])
    if parsed_values.get("app_user_id"):
        return build_borrower_query_graph(rows, parsed_values["app_user_id"])
    return {"type": "", "nodes": [], "edges": [], "summary": {"node_count": 0, "edge_count": 0, "isolated_count": 0}}


def build_agent_query_graph(rows: list[dict[str, Any]], input_agents: set[str]) -> dict[str, Any]:
    grouped: dict[str, dict[str, set[str]]] = defaultdict(lambda: defaultdict(set))
    for raw_row in rows:
        row = apply_aliases(raw_row)
        agent_id = normalize_phone(get_field(row, AGENT_COL))
        if agent_id not in input_agents:
            continue
        borrower_id = normalize_phone(get_field(row, BORROWER_COL))
        device_id = normalize_value(get_field(row, DEVICE_COL))
        ip = normalize_value(get_field(row, IP_COL))
        addr = normalize_value(get_field(row, ADDR_CLUSTER_COL))
        if borrower_id:
            grouped[agent_id]["users"].add(borrower_id)
        if device_id:
            grouped[agent_id]["devices"].add(device_id)
        if ip:
            grouped[agent_id]["ips"].add(ip)
        if addr:
            grouped[agent_id]["addrs"].add(addr)

    nodes = sorted(grouped)
    if len(nodes) > QUERY_GRAPH_MAX_NODES:
        return query_graph_too_large("agent", len(nodes))
    edges: list[dict[str, Any]] = []
    for left, right in combinations(nodes, 2):
        shared_users = grouped[left]["users"] & grouped[right]["users"]
        shared_devices = grouped[left]["devices"] & grouped[right]["devices"]
        shared_ips = grouped[left]["ips"] & grouped[right]["ips"]
        shared_addrs = grouped[left]["addrs"] & grouped[right]["addrs"]
        if not (shared_users or shared_devices or shared_ips or shared_addrs):
            continue
        edges.append(
            {
                "source": left,
                "target": right,
                "shared_user_count": len(shared_users),
                "shared_device_count": len(shared_devices),
                "shared_ip_count": len(shared_ips),
                "shared_addr_count": len(shared_addrs),
                "shared_users": sorted(shared_users),
                "shared_devices": sorted(shared_devices),
                "shared_ips": sorted(shared_ips),
                "shared_addrs": sorted(shared_addrs),
            }
        )
    return query_graph_payload("agent", nodes, edges)


def build_borrower_query_graph(rows: list[dict[str, Any]], input_borrowers: set[str]) -> dict[str, Any]:
    grouped: dict[str, dict[str, set[str]]] = defaultdict(lambda: defaultdict(set))
    for raw_row in rows:
        row = apply_aliases(raw_row)
        borrower_id = normalize_phone(get_field(row, BORROWER_COL))
        if borrower_id not in input_borrowers:
            continue
        agent_id = normalize_phone(get_field(row, AGENT_COL))
        device_id = normalize_value(get_field(row, DEVICE_COL))
        ip = normalize_value(get_field(row, IP_COL))
        addr = normalize_value(get_field(row, ADDR_CLUSTER_COL))
        if agent_id:
            grouped[borrower_id]["agents"].add(agent_id)
        if device_id:
            grouped[borrower_id]["devices"].add(device_id)
        if ip:
            grouped[borrower_id]["ips"].add(ip)
        if addr:
            grouped[borrower_id]["addrs"].add(addr)

    nodes = sorted(grouped)
    if len(nodes) > QUERY_GRAPH_MAX_NODES:
        return query_graph_too_large("borrower", len(nodes))
    edges: list[dict[str, Any]] = []
    for left, right in combinations(nodes, 2):
        shared_agents = grouped[left]["agents"] & grouped[right]["agents"]
        shared_devices = grouped[left]["devices"] & grouped[right]["devices"]
        shared_ips = grouped[left]["ips"] & grouped[right]["ips"]
        shared_addrs = grouped[left]["addrs"] & grouped[right]["addrs"]
        if not (shared_agents or shared_devices or shared_ips or shared_addrs):
            continue
        edges.append(
            {
                "source": left,
                "target": right,
                "shared_agent_count": len(shared_agents),
                "shared_device_count": len(shared_devices),
                "shared_ip_count": len(shared_ips),
                "shared_addr_count": len(shared_addrs),
                "shared_agents": sorted(shared_agents),
                "shared_devices": sorted(shared_devices),
                "shared_ips": sorted(shared_ips),
                "shared_addrs": sorted(shared_addrs),
            }
        )
    return query_graph_payload("borrower", nodes, edges)


def query_graph_payload(graph_type: str, node_ids: list[str], edges: list[dict[str, Any]]) -> dict[str, Any]:
    edge_nodes = {edge["source"] for edge in edges} | {edge["target"] for edge in edges}
    nodes = [
        {"id": node_id, "label": mask_phone(node_id) if graph_type in {"agent", "borrower"} else node_id}
        for node_id in node_ids
    ]
    return {
        "type": graph_type,
        "nodes": nodes,
        "edges": edges,
        "summary": {
            "node_count": len(nodes),
            "edge_count": len(edges),
            "isolated_count": sum(1 for node_id in node_ids if node_id not in edge_nodes),
        },
    }


def query_graph_too_large(graph_type: str, node_count: int) -> dict[str, Any]:
    return {
        "type": graph_type,
        "nodes": [],
        "edges": [],
        "too_many": True,
        "max_nodes": QUERY_GRAPH_MAX_NODES,
        "summary": {"node_count": node_count, "edge_count": 0, "isolated_count": 0},
    }


def summary_from_feature_table(feature_table: dict[str, Any]) -> dict[str, Any]:
    columns = feature_table.get("columns") or []
    rows = feature_table.get("rows") or []
    if "社区id(Agent)" not in columns:
        return {}

    community_ids: set[str] = set()
    max_community_agents = 0
    has_size_column = "社区规模(Agent)" in columns
    for row in rows:
        community_id = normalize_value(row.get("社区id(Agent)"))
        if not community_id:
            continue
        community_ids.add(community_id)
        if has_size_column:
            try:
                size = int(float(normalize_value(row.get("社区规模(Agent)")) or 0))
            except ValueError:
                size = 0
            max_community_agents = max(max_community_agents, size)

    return {
        "community_count": len(community_ids),
        "max_community_agents": max_community_agents,
        "max_community_size": max_community_agents,
        "community_size_label": "最大团伙中介数",
    }


def update_cluster_job(job_id: str | None, **updates: Any) -> None:
    if not job_id:
        return
    with CLUSTER_JOBS_LOCK:
        if job_id in CLUSTER_JOBS:
            CLUSTER_JOBS[job_id].update(updates)


def cluster_job(job_id: str) -> dict[str, Any]:
    with CLUSTER_JOBS_LOCK:
        if job_id not in CLUSTER_JOBS:
            raise ValueError("聚类任务不存在。")
        job = CLUSTER_JOBS[job_id]
        return {
            key: value
            for key, value in job.items()
            if key not in {"result"}
        }


def cluster_result(job_id: str) -> dict[str, Any]:
    with CLUSTER_JOBS_LOCK:
        if job_id not in CLUSTER_JOBS:
            raise ValueError("聚类任务不存在。")
        job = CLUSTER_JOBS[job_id]
        if job.get("status") != "done":
            raise ValueError("聚类任务尚未完成。")
        return job.get("result") or {}


def start_cluster_job(kind: str, target: Any, *args: Any) -> str:
    job_id = uuid.uuid4().hex
    with CLUSTER_JOBS_LOCK:
        CLUSTER_JOBS[job_id] = {
            "job_id": job_id,
            "kind": kind,
            "status": "pending",
            "progress": 0,
            "message": "等待开始",
        }

    def runner() -> None:
        try:
            update_cluster_job(job_id, status="running", progress=3, message="任务启动")
            result = target(job_id, *args)
            with CLUSTER_JOBS_LOCK:
                CLUSTER_JOBS[job_id].update(
                    {
                        "status": "done",
                        "progress": 100,
                        "message": "聚类完成",
                        "result": result,
                    }
                )
        except Exception as exc:  # noqa: BLE001 - surface task errors to UI
            detail = traceback.format_exc(limit=6)
            message = short_text(str(exc), 900)
            update_cluster_job(
                job_id,
                status="error",
                progress=100,
                message=message,
                error=message,
                error_detail=short_text(detail, 1600),
            )

    threading.Thread(target=runner, daemon=True).start()
    return job_id


def run_full_cluster_job(job_id: str, rows: list[dict[str, Any]], params: dict[str, Any]) -> dict[str, Any]:
    require_clustering_dependencies()
    if np is None or DBSCAN is None:
        raise ValueError("当前环境缺少聚类依赖。")
    eps = float(params.get("eps", 0.68))
    min_samples = int(params.get("min_samples", 1))
    metric = normalize_value(params.get("metric") or "euclidean")
    normalize_embeddings = bool(params.get("normalize", True))
    prepared_rows = rows_with_receiver_addr(rows)
    addresses = [row.get("receiverAddr") or "" for row in prepared_rows]
    if not addresses:
        raise ValueError("没有可聚类的地址数据。")
    update_cluster_job(job_id, progress=8, message=f"已读取 {len(addresses)} 条地址，准备编码")
    embeddings = embed_addresses(addresses, job_id)
    vectors = sklearn_normalize(embeddings) if normalize_embeddings and sklearn_normalize else embeddings
    update_cluster_job(job_id, progress=66, message="正在运行 DBSCAN")
    dbscan = DBSCAN(eps=eps, min_samples=min_samples, metric=metric)
    labels = [int(value) for value in dbscan.fit_predict(vectors)]
    update_cluster_job(job_id, progress=84, message="正在生成结果表")
    run_params = {
        "mode": "full",
        "model": EMBEDDING_MODEL_NAME,
        "eps": eps,
        "min_samples": min_samples,
        "metric": metric,
        "normalize": normalize_embeddings,
    }
    tables = build_cluster_tables(prepared_rows, labels)
    payload = engineering_payload(prepared_rows, labels, embeddings, run_params)
    tables.update(
        {
            "params": run_params,
            "stats": {
                "row_count": len(prepared_rows),
                "cluster_count": len({label for label in labels if label >= 0}),
                "noise_count": sum(1 for label in labels if label < 0),
            },
            "engineering_file": payload,
        }
    )
    return tables


def parse_engineering_file(payload: dict[str, Any]) -> tuple[list[str], list[int], Any]:
    if np is None:
        raise ValueError("当前环境缺少 numpy。")
    if not isinstance(payload, dict):
        raise ValueError(f"工程文件格式不正确：根对象应为 JSON object，实际为 {type(payload).__name__}。")
    items = payload.get("items") or []
    if not isinstance(items, list):
        raise ValueError("工程文件格式不正确：items 应为数组。请上传“导出工程文件”生成的 JSON，而不是明细表 CSV/Excel。")
    addresses: list[str] = []
    labels: list[int] = []
    vectors: list[Any] = []
    skipped_missing_address = 0
    skipped_missing_embedding = 0
    skipped_bad_embedding = 0
    sample_keys: list[str] = []
    for item in items:
        if not isinstance(item, dict):
            skipped_bad_embedding += 1
            continue
        if not sample_keys:
            sample_keys = list(item.keys())
        address = clean_receiver_address(item.get("receiverAddr"))
        vector = item.get("embedding")
        if not address:
            skipped_missing_address += 1
            continue
        if vector is None:
            skipped_missing_embedding += 1
            continue
        try:
            vector_values = np.asarray(vector, dtype="float32").tolist()
        except Exception:
            skipped_bad_embedding += 1
            continue
        addresses.append(address)
        labels.append(int(item.get("addr_cluster_id", -1)))
        EMBEDDING_CACHE.setdefault(address, vector_values)
        vectors.append(vector_values)
    if not addresses or not vectors:
        top_keys = list(payload.keys())[:12]
        raise ValueError(
            "工程文件缺少可用的 receiverAddr / addr_cluster_id / embedding。"
            f"已读取根字段：{top_keys or '无'}；items 数量：{len(items)}；"
            f"首个 item 字段：{sample_keys or '无'}；"
            f"缺地址：{skipped_missing_address}；缺 embedding：{skipped_missing_embedding}；embedding 无法解析：{skipped_bad_embedding}。"
            "请确认上传的是聚类汇总区“导出工程文件”生成的 JSON，不是明细表或汇总表。"
        )
    return addresses, labels, np.asarray(vectors, dtype="float32")


def run_faiss_incremental_assignment(
    job_id: str,
    base_vectors: Any,
    base_labels: list[int],
    new_vectors: Any,
    threshold: float,
    index_type: str,
    params: dict[str, Any],
) -> tuple[list[int], list[float]]:
    if np is None:
        raise ValueError("当前环境缺少 numpy。")
    worker_path = ROOT / "faiss_incremental_worker.py"
    if not worker_path.exists():
        raise ValueError("缺少 FAISS 增量 worker 文件：faiss_incremental_worker.py")
    with tempfile.TemporaryDirectory(prefix="anti_fraud_faiss_") as temp_dir:
        temp_path = Path(temp_dir)
        input_path = temp_path / "input.npz"
        output_path = temp_path / "output.json"
        np.savez_compressed(
            input_path,
            base_vectors=np.asarray(base_vectors, dtype="float32"),
            base_labels=np.asarray(base_labels, dtype="int64"),
            new_vectors=np.asarray(new_vectors, dtype="float32"),
            threshold=np.asarray([threshold], dtype="float32"),
            index_type=np.asarray([index_type]),
            params=np.asarray([json.dumps(params, ensure_ascii=False)]),
        )
        process = subprocess.Popen(
            [sys.executable, str(worker_path), str(input_path), str(output_path)],
            cwd=str(ROOT),
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            bufsize=1,
        )
        stderr_lines: list[str] = []
        if process.stdout is not None:
            for line in process.stdout:
                line = line.strip()
                if not line:
                    continue
                try:
                    message = json.loads(line)
                except json.JSONDecodeError:
                    continue
                message_type = message.get("type")
                if message_type == "progress":
                    done = int(message.get("done") or 0)
                    total = max(1, int(message.get("total") or 1))
                    progress = 64 + int(done / total * 24)
                    update_cluster_job(job_id, progress=progress, message=f"正在顺序归类：{done}/{total}")
                elif message_type == "error":
                    stderr_lines.append(str(message.get("error") or ""))
                    if message.get("traceback"):
                        stderr_lines.append(str(message.get("traceback")))
        stderr = ""
        if process.stderr is not None:
            stderr = process.stderr.read()
        exitcode = process.wait()
        if stderr:
            stderr_lines.append(stderr)
        if exitcode != 0:
            detail = short_text(" ".join(part for part in stderr_lines if part), 1200)
            raise ValueError(
                f"FAISS {index_type} 独立进程执行失败，exitcode={exitcode}。"
                f"{'原因：' + detail if detail else '没有返回详细错误。'}"
            )
        if not output_path.exists():
            raise ValueError(f"FAISS {index_type} 独立进程没有生成结果文件。")
        result = json.loads(output_path.read_text(encoding="utf-8"))
    return list(result.get("assigned") or []), list(result.get("distances") or [])


def run_incremental_cluster_job(
    job_id: str,
    base_payload: dict[str, Any],
    incremental_rows: list[dict[str, Any]],
    params: dict[str, Any],
) -> dict[str, Any]:
    require_clustering_dependencies(incremental=True)
    if np is None:
        raise ValueError("当前环境缺少 numpy。")
    threshold = float(params.get("threshold", 0.68))
    index_type = normalize_value(params.get("index_type") or "FlatL2")
    normalize_embeddings = bool(params.get("normalize", True))
    base_addresses, base_labels, base_embeddings = parse_engineering_file(base_payload)
    prepared_rows = rows_with_receiver_addr(incremental_rows)
    addresses = [row.get("receiverAddr") or "" for row in prepared_rows]
    if not addresses:
        raise ValueError("没有可增量聚类的地址数据。")
    update_cluster_job(job_id, progress=8, message="正在编码增量地址")
    new_embeddings = embed_addresses(addresses, job_id)
    update_cluster_job(job_id, progress=60, message="增量地址编码完成，正在准备向量")
    base_vectors = sklearn_normalize(base_embeddings) if normalize_embeddings and sklearn_normalize else base_embeddings
    new_vectors = sklearn_normalize(new_embeddings) if normalize_embeddings and sklearn_normalize else new_embeddings
    update_cluster_job(job_id, progress=62, message=f"正在使用 FAISS {index_type} 索引执行增量归类")
    assigned, distances = run_faiss_incremental_assignment(
        job_id,
        base_vectors.astype("float32"),
        base_labels,
        new_vectors.astype("float32"),
        threshold,
        index_type,
        params,
    )
    update_cluster_job(job_id, progress=90, message="正在生成增量结果表")
    run_params = {
        "mode": "incremental",
        "model": EMBEDDING_MODEL_NAME,
        "index_type": index_type,
        "threshold": threshold,
        "normalize": normalize_embeddings,
        "hnsw_m": int(params.get("hnsw_m", 32)),
        "hnsw_ef_search": int(params.get("hnsw_ef_search", 64)),
        "ivf_nlist": int(params.get("ivf_nlist", 64)),
        "ivf_nprobe": int(params.get("ivf_nprobe", 8)),
    }
    tables = build_cluster_tables(prepared_rows, assigned)
    combined_addresses = base_addresses + addresses
    combined_labels = base_labels + assigned
    combined_embeddings = np.vstack([base_embeddings, new_embeddings]).astype("float32")
    combined_rows = [{"receiverAddr": address} for address in combined_addresses]
    payload = engineering_payload(combined_rows, combined_labels, combined_embeddings, run_params)
    detail_rows = tables["detail_table"]["rows"]
    for row, distance in zip(detail_rows, distances):
        row["nearest_distance"] = distance
    if "nearest_distance" not in tables["detail_table"]["columns"]:
        tables["detail_table"]["columns"].append("nearest_distance")
    tables.update(
        {
            "params": run_params,
            "stats": {
                "row_count": len(prepared_rows),
                "cluster_count": len(set(assigned)),
                "new_cluster_count": sum(1 for label in set(assigned) if label >= max(base_labels, default=-1) + 1),
            },
            "engineering_file": payload,
        }
    )
    return tables


def as_boolish(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return False
        return value > 0
    text = str(value).strip().lower()
    if not text or text in {"0", "0.0", "false", "否", "无", "none", "null", "nan", "n", "no"}:
        return False
    if text in {"1", "1.0", "true", "是", "有", "逾期", "y", "yes"}:
        return True
    try:
        return float(text) > 0
    except ValueError:
        return bool(text)


def mask_phone(value: str) -> str:
    if len(value) >= 7 and value.isdigit():
        return f"{value[:3]}****{value[-4:]}"
    if len(value) > 8:
        return f"{value[:4]}...{value[-4:]}"
    return value


def read_csv_rows(content: bytes) -> list[dict[str, Any]]:
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "gb18030", "utf-16"):
        try:
            text = content.decode(encoding)
            reader = csv.DictReader(io.StringIO(text))
            reader.fieldnames = [normalize_header(name) for name in (reader.fieldnames or [])]
            return [{normalize_header(k): json_safe_value(v) for k, v in row.items()} for row in reader]
        except Exception as exc:  # noqa: BLE001 - try several common spreadsheet encodings
            last_error = exc
    raise ValueError(f"CSV 解析失败：{last_error}")


def read_excel_rows(content: bytes) -> list[dict[str, Any]]:
    if load_workbook is None:
        raise ValueError("当前 Python 环境缺少 openpyxl，无法解析 Excel 文件。")
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = [normalize_header(v) for v in next(rows)]
    except StopIteration as exc:
        raise ValueError("Excel 文件为空。") from exc
    records: list[dict[str, Any]] = []
    for row in rows:
        if not any(cell not in (None, "") for cell in row):
            continue
        records.append({headers[index]: json_safe_value(value) for index, value in enumerate(row) if index < len(headers)})
    return records


def parse_uploaded_file(filename: str, content_base64: str) -> list[dict[str, Any]]:
    content = base64.b64decode(content_base64)
    suffix = Path(filename).suffix.lower()
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return read_excel_rows(content)
    if suffix == ".xls":
        raise ValueError("暂不支持旧版 .xls，请另存为 .xlsx 或 CSV 后上传。")
    return read_csv_rows(content)


class RiskGraph:
    def __init__(self, rows: list[dict[str, Any]], overdue_basis: str = "any", reloan_filter: str = "all"):
        self.rows = rows
        self.overdue_basis = overdue_basis
        self.reloan_filter = normalize_value(reloan_filter or "all")
        self.agent_to_borrowers: dict[str, set[str]] = defaultdict(set)
        self.agent_direct_attributes: dict[str, dict[str, set[str]]] = defaultdict(lambda: defaultdict(set))
        self.borrower_to_agents: dict[str, set[str]] = defaultdict(set)
        self.loans_by_borrower: dict[str, list[dict[str, Any]]] = defaultdict(list)
        self.loan_rows: list[dict[str, Any]] = []
        self.skipped_rows = 0
        self.source_columns = list(rows[0].keys()) if rows else []
        self.columns = sorted({column for row in rows for column in row})
        self._agent_relation_counts_cache: dict[int, dict[tuple[str, str], dict[str, int]]] = {}
        self._build()

    def _build(self) -> None:
        for row in self.rows:
            row = apply_aliases(row)
            agent_id = normalize_phone(get_field(row, AGENT_COL))
            borrower_id = normalize_phone(get_field(row, BORROWER_COL))
            if not agent_id or not borrower_id:
                self.skipped_rows += 1
                continue
            clean_row = {key: json_safe_value(value) for key, value in row.items()}
            clean_row[AGENT_COL] = agent_id
            clean_row[BORROWER_COL] = borrower_id
            clean_row[DEVICE_COL] = normalize_value(get_field(row, DEVICE_COL))
            clean_row[IP_COL] = normalize_value(get_field(row, IP_COL))
            clean_row[ADDRESS_COL] = normalize_value(get_field(row, ADDRESS_COL))
            clean_row[ADDR_CLUSTER_COL] = normalize_value(get_field(row, ADDR_CLUSTER_COL))
            clean_row[RELOAN_COL] = normalize_value(get_field(row, RELOAN_COL))
            clean_row[RETURN_COL] = normalize_value(get_field(row, RETURN_COL))
            self.agent_to_borrowers[agent_id].add(borrower_id)
            self.borrower_to_agents[borrower_id].add(agent_id)
            self.loans_by_borrower[borrower_id].append(clean_row)
            self.loan_rows.append(clean_row)
            for column in (DEVICE_COL, IP_COL, ADDR_CLUSTER_COL):
                value = normalize_value(clean_row.get(column))
                if value:
                    self.agent_direct_attributes[agent_id][column].add(value)

    def distinct_from_loans(self, loans: list[dict[str, Any]], column: str) -> set[str]:
        return {normalize_value(loan.get(column)) for loan in loans if normalize_value(loan.get(column))}

    def loan_group_stats(self, loans: list[dict[str, Any]], basis: str | None = None) -> dict[str, Any]:
        funded = self.funded_loans(loans)
        overdue = sum(1 for loan in funded if self.is_overdue(loan, basis))
        return {
            "borrowers": self.distinct_from_loans(loans, BORROWER_COL),
            "agents": self.distinct_from_loans(loans, AGENT_COL),
            "devices": self.distinct_from_loans(loans, DEVICE_COL),
            "ips": self.distinct_from_loans(loans, IP_COL),
            "addr_clusters": self.distinct_from_loans(loans, ADDR_CLUSTER_COL),
            "addresses": self.distinct_from_loans(loans, ADDRESS_COL),
            "application_count": len(loans),
            "loan_count": len(funded),
            "overdue_loan_count": overdue,
            "bad_rate": overdue / len(funded) if funded else 0.0,
        }

    def bad_rate(self, loans: list[dict[str, Any]], basis: str | None = None) -> float:
        funded = self.funded_loans(loans)
        return sum(1 for loan in funded if self.is_overdue(loan, basis)) / len(funded) if funded else 0.0

    def is_funded(self, loan: dict[str, Any]) -> bool:
        return normalize_value(loan.get(FINAL_RESULT_COL)) == FUNDED_RESULT

    def funded_loans(self, loans: list[dict[str, Any]]) -> list[dict[str, Any]]:
        return [loan for loan in loans if self.is_funded(loan)]

    def loans_for_borrowers(self, borrowers: set[str]) -> list[dict[str, Any]]:
        return [loan for borrower_id in borrowers for loan in self.loans_by_borrower.get(borrower_id, [])]

    def loans_where(self, column: str, value: str) -> list[dict[str, Any]]:
        if not value:
            return []
        return [loan for loan in self.loan_rows if normalize_value(loan.get(column)) == value]

    def first_degree_borrowers(self, agent_id: str) -> set[str]:
        return set(self.agent_to_borrowers.get(agent_id, set()))

    def first_degree_agents(self, borrower_id: str) -> set[str]:
        return set(self.borrower_to_agents.get(borrower_id, set()))

    def second_degree_agents(self, agent_id: str) -> set[str]:
        result: set[str] = set()
        for borrower_id in self.first_degree_borrowers(agent_id):
            result.update(self.borrower_to_agents.get(borrower_id, set()))
        result.discard(agent_id)
        return result

    def second_degree_borrowers(self, borrower_id: str) -> set[str]:
        result: set[str] = set()
        for agent_id in self.first_degree_agents(borrower_id):
            result.update(self.agent_to_borrowers.get(agent_id, set()))
        result.discard(borrower_id)
        return result

    def is_overdue(self, loan: dict[str, Any], basis: str | None = None) -> bool:
        field = basis or self.overdue_basis
        if field == "any":
            return any(as_boolish(loan.get(name)) for name in OVERDUE_FIELDS)
        if field not in OVERDUE_FIELDS:
            field = "any"
        return as_boolish(loan.get(field))

    def second_degree_overdue_rate(self, borrower_id: str, basis: str | None = None) -> dict[str, Any]:
        borrowers = self.second_degree_borrowers(borrower_id)
        loans = self.funded_loans([loan for item in borrowers for loan in self.loans_by_borrower.get(item, [])])
        total = len(loans)
        overdue = sum(1 for loan in loans if self.is_overdue(loan, basis))
        return {
            "second_degree_borrowers": len(borrowers),
            "second_degree_loans": total,
            "second_degree_overdue_loans": overdue,
            "second_degree_overdue_rate": overdue / total if total else 0,
        }

    def agent_metrics(self, agent_id: str) -> dict[str, Any]:
        first = self.first_degree_borrowers(agent_id)
        second = self.second_degree_agents(agent_id)
        return {
            "agent_id": agent_id,
            "label": mask_phone(agent_id),
            "first_degree_borrowers": len(first),
            "second_degree_agents": len(second),
        }

    def borrower_metrics(self, borrower_id: str, basis: str | None = None) -> dict[str, Any]:
        first = self.first_degree_agents(borrower_id)
        rate = self.second_degree_overdue_rate(borrower_id, basis)
        own_loans = self.funded_loans(self.loans_by_borrower.get(borrower_id, []))
        overdue_loans = sum(1 for loan in own_loans if self.is_overdue(loan, basis))
        return {
            "borrower_id": borrower_id,
            "label": mask_phone(borrower_id),
            "first_degree_agents": len(first),
            "loan_count": len(own_loans),
            "own_overdue_loans": overdue_loans,
            **rate,
        }

    def top_agents(self, limit: int = 20) -> list[dict[str, Any]]:
        agents = [self.agent_metrics(agent_id) for agent_id in self.agent_to_borrowers]
        return sorted(agents, key=lambda item: (item["second_degree_agents"], item["first_degree_borrowers"]), reverse=True)[:limit]

    def top_borrowers(self, basis: str | None = None, limit: int = 20) -> list[dict[str, Any]]:
        borrowers = [self.borrower_metrics(borrower_id, basis) for borrower_id in self.borrower_to_agents]
        return sorted(
            borrowers,
            key=lambda item: (
                item["second_degree_overdue_rate"],
                item["second_degree_borrowers"],
                item["second_degree_loans"],
            ),
            reverse=True,
        )[:limit]

    def agent_projection(self, min_shared_borrowers: int = 1) -> dict[str, dict[str, int]]:
        raw_edges: dict[tuple[str, str], int] = defaultdict(int)
        for agents in self.borrower_to_agents.values():
            for left, right in combinations(sorted(agents), 2):
                raw_edges[(left, right)] += 1

        adjacency: dict[str, dict[str, int]] = defaultdict(dict)
        for (left, right), weight in raw_edges.items():
            if weight < min_shared_borrowers:
                continue
            adjacency[left][right] = weight
            adjacency[right][left] = weight
        return adjacency

    def agent_attribute_sets(self, column: str) -> dict[str, set[str]]:
        if column == BORROWER_COL:
            return {agent_id: set(borrowers) for agent_id, borrowers in self.agent_to_borrowers.items()}
        return {agent_id: set(self.agent_direct_attributes.get(agent_id, {}).get(column, set())) for agent_id in self.agent_to_borrowers}

    def agent_direct_values(self, agent_id: str, column: str) -> set[str]:
        return set(self.agent_direct_attributes.get(agent_id, {}).get(column, set()))

    def shared_agent_values(self, left: str, right: str, column: str) -> set[str]:
        return self.agent_direct_values(left, column) & self.agent_direct_values(right, column)

    def connected_component_rule(self) -> dict[str, int]:
        if self.reloan_filter == "all":
            return {"shared_user": 2, "ip_borrower_limit": 50}
        return {"shared_user": 1, "ip_borrower_limit": 20}

    def agent_relation_counts(self, ip_borrower_limit: int | None = None) -> dict[tuple[str, str], dict[str, int]]:
        if ip_borrower_limit is None:
            ip_borrower_limit = self.connected_component_rule()["ip_borrower_limit"]
        if ip_borrower_limit in self._agent_relation_counts_cache:
            return self._agent_relation_counts_cache[ip_borrower_limit]
        agents = sorted(self.agent_to_borrowers)
        borrower_sets = self.agent_attribute_sets(BORROWER_COL)
        device_sets = self.agent_attribute_sets(DEVICE_COL)
        ip_sets = self.agent_attribute_sets(IP_COL)
        addr_sets = self.agent_attribute_sets(ADDR_CLUSTER_COL)
        ip_to_borrowers: dict[str, set[str]] = defaultdict(set)
        for loan in self.loan_rows:
            ip = normalize_value(loan.get(IP_COL))
            borrower_id = normalize_phone(loan.get(BORROWER_COL))
            if ip and borrower_id:
                ip_to_borrowers[ip].add(borrower_id)
        eligible_ips = {ip for ip, borrowers in ip_to_borrowers.items() if len(borrowers) <= ip_borrower_limit}

        counts: dict[tuple[str, str], dict[str, int]] = {}
        for left, right in combinations(agents, 2):
            shared_user = len(borrower_sets.get(left, set()) & borrower_sets.get(right, set()))
            shared_device_values = device_sets.get(left, set()) & device_sets.get(right, set())
            shared_ip_all = ip_sets.get(left, set()) & ip_sets.get(right, set())
            shared_ip = len(shared_ip_all & eligible_ips)
            shared_addr_values = addr_sets.get(left, set()) & addr_sets.get(right, set())
            if shared_user or shared_device_values or shared_ip_all or shared_addr_values:
                counts[(left, right)] = {
                    "user": shared_user,
                    "device": len(shared_device_values),
                    "ip": shared_ip,
                    "ip_all": len(shared_ip_all),
                    "address": len(shared_addr_values),
                }
        self._agent_relation_counts_cache[ip_borrower_limit] = counts
        return counts

    def filtered_agent_projection(self) -> dict[str, dict[str, float]]:
        rule = self.connected_component_rule()
        adjacency: dict[str, dict[str, float]] = defaultdict(dict)
        for (left, right), counts in self.agent_relation_counts(rule["ip_borrower_limit"]).items():
            if not (
                counts["device"] >= 1
                or counts["user"] >= rule["shared_user"]
                or (counts["ip"] >= 2)
                or counts["address"] >= 1
            ):
                continue
            weight = counts["user"] + counts["device"] * 1.2 + counts["ip"] * 0.8 + counts["address"]
            adjacency[left][right] = weight
            adjacency[right][left] = weight
        return adjacency

    def agent_multi_projection(self) -> dict[str, dict[str, float]]:
        adjacency: dict[str, dict[str, float]] = defaultdict(dict)
        for (left, right), counts in self.agent_relation_counts().items():
            weight = counts["user"] + counts["device"] * 1.2 + counts["ip_all"] * 0.8 + counts["address"]
            if not weight:
                continue
            adjacency[left][right] = weight
            adjacency[right][left] = weight
        for agent_id in self.agent_to_borrowers:
            adjacency.setdefault(agent_id, {})
        return adjacency

    def agent_projection_by_column(self, column: str) -> dict[str, dict[str, int]]:
        adjacency: dict[str, dict[str, int]] = defaultdict(dict)
        attribute_sets = self.agent_attribute_sets(column)
        for left, right in combinations(sorted(self.agent_to_borrowers), 2):
            weight = len(attribute_sets.get(left, set()) & attribute_sets.get(right, set()))
            if not weight:
                continue
            adjacency[left][right] = weight
            adjacency[right][left] = weight
        return adjacency

    def agent_components_from_adjacency(
        self,
        adjacency: dict[str, dict[str, float]] | dict[str, dict[str, int]],
        min_agents: int = 1,
    ) -> list[set[str]]:
        seen: set[str] = set()
        components: list[set[str]] = []
        for agent_id in sorted(self.agent_to_borrowers):
            if agent_id in seen:
                continue
            stack = [agent_id]
            component: set[str] = set()
            seen.add(agent_id)
            while stack:
                current = stack.pop()
                component.add(current)
                for neighbor in adjacency.get(current, {}):
                    if neighbor not in seen:
                        seen.add(neighbor)
                        stack.append(neighbor)
            if len(component) >= min_agents:
                components.append(component)
        return components

    def pagerank(
        self,
        adjacency: dict[str, dict[str, float]] | dict[str, dict[str, int]],
        damping: float = 0.85,
        iterations: int = 40,
    ) -> dict[str, float]:
        nodes = sorted(self.agent_to_borrowers)
        if not nodes:
            return {}
        n = len(nodes)
        node_index = {node: index for index, node in enumerate(nodes)}
        scores = [1 / n] * n
        transitions: list[list[tuple[int, float]]] = []
        dangling_indices: list[int] = []
        for node in nodes:
            neighbors = adjacency.get(node, {})
            weight_sum = sum(neighbors.values())
            if not weight_sum:
                transitions.append([])
                dangling_indices.append(node_index[node])
                continue
            transitions.append(
                [
                    (node_index[neighbor], float(weight) / weight_sum)
                    for neighbor, weight in neighbors.items()
                    if neighbor in node_index
                ]
            )
        base_score = (1 - damping) / n
        for _ in range(iterations):
            dangling = sum(scores[index] for index in dangling_indices)
            dangling_share = damping * dangling / n
            next_scores = [base_score + dangling_share] * n
            for index, neighbors in enumerate(transitions):
                if not neighbors:
                    continue
                contribution = damping * scores[index]
                for neighbor_index, probability in neighbors:
                    next_scores[neighbor_index] += contribution * probability
            scores = next_scores
        return dict(zip(nodes, scores))

    def user_projection(self) -> dict[str, dict[str, float]]:
        value_to_borrowers: dict[tuple[str, str], set[str]] = defaultdict(set)
        relation_columns = (AGENT_COL, DEVICE_COL, IP_COL, ADDR_CLUSTER_COL)
        relation_weights = {
            AGENT_COL: 1.0,
            DEVICE_COL: 1.2,
            IP_COL: 0.8,
            ADDR_CLUSTER_COL: 1.0,
        }
        for loan in self.loan_rows:
            borrower_id = normalize_phone(loan.get(BORROWER_COL))
            if not borrower_id:
                continue
            for column in relation_columns:
                value = normalize_value(loan.get(column))
                if value:
                    value_to_borrowers[(column, value)].add(borrower_id)

        raw_edges: dict[tuple[str, str], float] = defaultdict(float)
        for (column, _), borrowers in value_to_borrowers.items():
            weight = relation_weights.get(column, 1.0)
            for left, right in combinations(sorted(borrowers), 2):
                raw_edges[(left, right)] += weight

        adjacency: dict[str, dict[str, float]] = defaultdict(dict)
        for borrower_id in self.borrower_to_agents:
            adjacency.setdefault(borrower_id, {})
        for (left, right), weight in raw_edges.items():
            adjacency[left][right] = weight
            adjacency[right][left] = weight
        return adjacency

    def igraph_from_adjacency(
        self,
        adjacency: dict[str, dict[str, float]],
    ) -> tuple[Any, list[str]]:
        if ig is None:
            raise ValueError("当前环境缺少 igraph。请先执行：python3 -m pip install -r requirements.txt")
        nodes = sorted(adjacency)
        index = {node: position for position, node in enumerate(nodes)}
        edges: list[tuple[int, int]] = []
        weights: list[float] = []
        for left in nodes:
            for right, weight in adjacency.get(left, {}).items():
                if right not in index or left >= right:
                    continue
                edges.append((index[left], index[right]))
                weights.append(float(weight))
        graph = ig.Graph(n=len(nodes), edges=edges, directed=False)
        graph.vs["name"] = nodes
        graph.es["weight"] = weights
        return graph, nodes

    def partition_from_membership(self, nodes: list[str], membership: list[int], prefix: str) -> dict[str, str]:
        return self.normalize_partition_ids(
            {node: str(membership[index]) for index, node in enumerate(nodes)},
            prefix,
        )

    def louvain_partition(
        self,
        adjacency: dict[str, dict[str, float]],
        resolution: float = 1.0,
    ) -> dict[str, str]:
        graph, nodes = self.igraph_from_adjacency(adjacency)
        if not nodes:
            return {}
        if graph.ecount() == 0:
            return self.normalize_partition_ids({node: node for node in nodes}, "L")
        partition = graph.community_multilevel(weights="weight", resolution=resolution)
        return self.partition_from_membership(nodes, partition.membership, "L")

    def leiden_partition(
        self,
        adjacency: dict[str, dict[str, float]],
        resolution: float = 1.0,
    ) -> dict[str, str]:
        graph, nodes = self.igraph_from_adjacency(adjacency)
        if not nodes:
            return {}
        if graph.ecount() == 0:
            return self.normalize_partition_ids({node: node for node in nodes}, "LD")
        if leidenalg is None:
            raise ValueError("当前环境缺少 leidenalg。请先执行：python3 -m pip install -r requirements.txt")
        partition = leidenalg.find_partition(
            graph,
            leidenalg.RBConfigurationVertexPartition,
            weights="weight",
            resolution_parameter=resolution,
        )
        return self.partition_from_membership(nodes, partition.membership, "LD")

    def graph_quality_metrics(self, community_method: str, resolution: float = 1.0) -> dict[str, Any]:
        user_adjacency = self.user_projection()
        agent_adjacency = self.agent_multi_projection()
        metrics: dict[str, Any] = {}

        for prefix, adjacency in (("user", user_adjacency), ("agent", agent_adjacency)):
            graph, _ = self.igraph_from_adjacency(adjacency)
            metrics[f"{prefix}_graph_nodes"] = graph.vcount()
            metrics[f"{prefix}_graph_edges"] = graph.ecount()
            if not graph.vcount() or not graph.ecount():
                continue
            if community_method == "louvain":
                partition = graph.community_multilevel(weights="weight", resolution=resolution)
                metrics[f"{prefix}_louvain_q"] = graph.modularity(partition.membership, weights="weight")
            elif community_method == "leiden":
                if leidenalg is None:
                    raise ValueError("当前环境缺少 leidenalg。请先执行：python3 -m pip install -r requirements.txt")
                partition = leidenalg.find_partition(
                    graph,
                    leidenalg.RBConfigurationVertexPartition,
                    weights="weight",
                    resolution_parameter=resolution,
                )
                metrics[f"{prefix}_leiden_quality"] = partition.quality()
        return metrics

    def split_disconnected_partition(
        self,
        partition: dict[str, str],
        adjacency: dict[str, dict[str, float]],
    ) -> dict[str, str]:
        by_community: dict[str, set[str]] = defaultdict(set)
        for node, community_id in partition.items():
            by_community[community_id].add(node)

        refined: dict[str, str] = {}
        counter = 1
        for community_nodes in by_community.values():
            seen: set[str] = set()
            for start in sorted(community_nodes):
                if start in seen:
                    continue
                stack = [start]
                component: set[str] = set()
                seen.add(start)
                while stack:
                    current = stack.pop()
                    component.add(current)
                    for neighbor in adjacency.get(current, {}):
                        if neighbor in community_nodes and neighbor not in seen:
                            seen.add(neighbor)
                            stack.append(neighbor)
                community_id = f"R{counter:03d}"
                counter += 1
                for node in component:
                    refined[node] = community_id
        return refined

    def normalize_partition_ids(self, partition: dict[str, str], prefix: str) -> dict[str, str]:
        groups: dict[str, list[str]] = defaultdict(list)
        for node, community_id in partition.items():
            groups[community_id].append(node)
        ordered = sorted(groups.values(), key=lambda nodes: (-len(nodes), sorted(nodes)[0]))
        normalized: dict[str, str] = {}
        for index, nodes in enumerate(ordered, 1):
            community_id = f"{prefix}{index:03d}"
            for node in nodes:
                normalized[node] = community_id
        return normalized

    def agent_community_components(self, min_shared_borrowers: int = 1, min_agents: int = 2) -> list[set[str]]:
        adjacency = self.filtered_agent_projection()
        return self.agent_components_from_adjacency(adjacency, min_agents)

    def community_betweenness(self, agents: set[str], adjacency: dict[str, dict[str, float]]) -> dict[str, float]:
        # Exact Brandes centrality is useful for finding bridge-like intermediaries,
        # but it can be expensive on very large communities.
        if len(agents) > 80:
            return {agent_id: 0.0 for agent_id in agents}

        centrality = {agent_id: 0.0 for agent_id in agents}
        for source in agents:
            stack: list[str] = []
            predecessors = {agent_id: [] for agent_id in agents}
            sigma = dict.fromkeys(agents, 0.0)
            distance = dict.fromkeys(agents, -1)
            sigma[source] = 1.0
            distance[source] = 0
            queue = [source]
            for current in queue:
                stack.append(current)
                for neighbor in adjacency.get(current, {}):
                    if neighbor not in agents:
                        continue
                    if distance[neighbor] < 0:
                        queue.append(neighbor)
                        distance[neighbor] = distance[current] + 1
                    if distance[neighbor] == distance[current] + 1:
                        sigma[neighbor] += sigma[current]
                        predecessors[neighbor].append(current)

            dependency = dict.fromkeys(agents, 0.0)
            while stack:
                node = stack.pop()
                for predecessor in predecessors[node]:
                    if sigma[node]:
                        dependency[predecessor] += (sigma[predecessor] / sigma[node]) * (1 + dependency[node])
                if node != source:
                    centrality[node] += dependency[node]

        scale = (len(agents) - 1) * (len(agents) - 2)
        if scale > 0:
            centrality = {agent_id: value / scale for agent_id, value in centrality.items()}
        return centrality

    def community_metrics(
        self,
        community_id: str,
        agents: set[str],
        adjacency: dict[str, dict[str, float]],
        basis: str | None = None,
    ) -> dict[str, Any]:
        borrowers = {borrower_id for agent_id in agents for borrower_id in self.agent_to_borrowers.get(agent_id, set())}
        loans = self.funded_loans([loan for borrower_id in borrowers for loan in self.loans_by_borrower.get(borrower_id, [])])
        overdue = sum(1 for loan in loans if self.is_overdue(loan, basis))
        internal_edges: list[tuple[str, str, float]] = []
        for left in agents:
            for right, weight in adjacency.get(left, {}).items():
                if right in agents and left < right:
                    internal_edges.append((left, right, weight))

        agent_count = len(agents)
        borrower_count = len(borrowers)
        edge_count = len(internal_edges)
        possible_edges = agent_count * (agent_count - 1) / 2
        total_shared = sum(weight for _, _, weight in internal_edges)
        shared_borrowers = {
            borrower_id
            for borrower_id in borrowers
            if len(self.borrower_to_agents.get(borrower_id, set()) & agents) >= 2
        }
        possible_edges = agent_count * (agent_count - 1) / 2
        device_density = self.community_attribute_density(agents, DEVICE_COL, possible_edges)
        ip_density = self.community_attribute_density(agents, IP_COL, possible_edges)
        address_density = self.community_attribute_density(agents, ADDR_CLUSTER_COL, possible_edges)
        degree = {
            agent_id: sum(1 for neighbor in adjacency.get(agent_id, {}) if neighbor in agents)
            for agent_id in agents
        }
        weighted_degree = {
            agent_id: sum(weight for neighbor, weight in adjacency.get(agent_id, {}).items() if neighbor in agents)
            for agent_id in agents
        }
        betweenness = self.community_betweenness(agents, adjacency)
        central_agents = sorted(
            agents,
            key=lambda agent_id: (
                weighted_degree.get(agent_id, 0),
                degree.get(agent_id, 0),
                betweenness.get(agent_id, 0),
                len(self.agent_to_borrowers.get(agent_id, set())),
            ),
            reverse=True,
        )[:5]
        overdue_rate = overdue / len(loans) if loans else 0
        density = edge_count / possible_edges if possible_edges else 0
        shared_ratio = len(shared_borrowers) / borrower_count if borrower_count else 0
        risk_score = min(
            100,
            round(
                overdue_rate * 55
                + density * 20
                + min(shared_ratio, 1) * 15
                + min(agent_count / 10, 1) * 10,
                2,
            ),
        )

        return {
            "community_id": community_id,
            "community_type": "agent",
            "agent_count": agent_count,
            "borrower_count": borrower_count,
            "loan_count": len(loans),
            "overdue_loan_count": overdue,
            "bad_debt_rate": overdue_rate,
            "edge_count": edge_count,
            "density": density,
            "shared_borrower_count": len(shared_borrowers),
            "avg_shared_borrowers_per_edge": total_shared / edge_count if edge_count else 0,
            "avg_degree": sum(degree.values()) / agent_count if agent_count else 0,
            "max_degree": max(degree.values()) if degree else 0,
            "device_density": device_density,
            "ip_density": ip_density,
            "address_density": address_density,
            "risk_score": risk_score,
            "top_agents": [
                {
                    "agent_id": agent_id,
                    "label": mask_phone(agent_id),
                    "degree": degree.get(agent_id, 0),
                    "weighted_degree": weighted_degree.get(agent_id, 0),
                    "degree_centrality": degree.get(agent_id, 0) / (agent_count - 1) if agent_count > 1 else 0,
                    "betweenness_centrality": betweenness.get(agent_id, 0),
                    "first_degree_borrowers": len(self.agent_to_borrowers.get(agent_id, set())),
                }
                for agent_id in central_agents
            ],
        }

    def community_attribute_density(self, agents: set[str], column: str, possible_edges: float) -> float:
        if not possible_edges:
            return 0.0
        value_to_agents: dict[str, set[str]] = defaultdict(set)
        for agent_id in agents:
            for value in self.agent_direct_attributes.get(agent_id, {}).get(column, set()):
                value_to_agents[value].add(agent_id)
        edges: set[tuple[str, str]] = set()
        for linked_agents in value_to_agents.values():
            for left, right in combinations(sorted(linked_agents), 2):
                edges.add((left, right))
        return len(edges) / possible_edges

    def user_community_metrics(
        self,
        community_id: str,
        borrowers: set[str],
        adjacency: dict[str, dict[str, float]],
        basis: str | None = None,
    ) -> dict[str, Any]:
        loans = self.funded_loans(self.loans_for_borrowers(borrowers))
        all_loans = self.loans_for_borrowers(borrowers)
        overdue = sum(1 for loan in loans if self.is_overdue(loan, basis))
        agents = self.distinct_from_loans(all_loans, AGENT_COL)
        edge_count = sum(
            1
            for left in borrowers
            for right in adjacency.get(left, {})
            if right in borrowers and left < right
        )
        possible_edges = len(borrowers) * (len(borrowers) - 1) / 2
        density = edge_count / possible_edges if possible_edges else 0
        bad_rate = overdue / len(loans) if loans else 0
        device_density = self.user_attribute_density(borrowers, DEVICE_COL, possible_edges)
        ip_density = self.user_attribute_density(borrowers, IP_COL, possible_edges)
        address_density = self.user_attribute_density(borrowers, ADDR_CLUSTER_COL, possible_edges)
        risk_score = min(
            100,
            round(
                bad_rate * 55
                + density * 15
                + device_density * 10
                + ip_density * 10
                + address_density * 10,
                2,
            ),
        )
        top_agents = sorted(
            agents,
            key=lambda agent_id: len(self.agent_to_borrowers.get(agent_id, set()) & borrowers),
            reverse=True,
        )[:5]
        return {
            "community_id": community_id,
            "community_type": "user",
            "agent_count": len(agents),
            "borrower_count": len(borrowers),
            "loan_count": len(loans),
            "overdue_loan_count": overdue,
            "bad_debt_rate": bad_rate,
            "edge_count": edge_count,
            "density": density,
            "shared_borrower_count": 0,
            "avg_shared_borrowers_per_edge": 0,
            "avg_degree": (2 * edge_count / len(borrowers)) if borrowers else 0,
            "max_degree": max(
                (sum(1 for neighbor in adjacency.get(node, {}) if neighbor in borrowers) for node in borrowers),
                default=0,
            ),
            "device_density": device_density,
            "ip_density": ip_density,
            "address_density": address_density,
            "risk_score": risk_score,
            "top_agents": [
                {
                    "agent_id": agent_id,
                    "label": mask_phone(agent_id),
                    "degree": len(self.agent_to_borrowers.get(agent_id, set()) & borrowers),
                    "weighted_degree": len(self.agent_to_borrowers.get(agent_id, set()) & borrowers),
                    "degree_centrality": 0,
                    "betweenness_centrality": 0,
                    "first_degree_borrowers": len(self.agent_to_borrowers.get(agent_id, set())),
                }
                for agent_id in top_agents
            ],
        }

    def user_attribute_density(self, borrowers: set[str], column: str, possible_edges: float) -> float:
        if not possible_edges:
            return 0.0
        value_to_borrowers: dict[str, set[str]] = defaultdict(set)
        for borrower_id in borrowers:
            for loan in self.loans_by_borrower.get(borrower_id, []):
                value = normalize_value(loan.get(column))
                if value:
                    value_to_borrowers[value].add(borrower_id)
        edges: set[tuple[str, str]] = set()
        for linked_borrowers in value_to_borrowers.values():
            for left, right in combinations(sorted(linked_borrowers), 2):
                edges.add((left, right))
        return len(edges) / possible_edges

    def user_communities(self, method: str, basis: str | None = None, limit: int = 20) -> list[dict[str, Any]]:
        adjacency = self.user_projection()
        partition = self.leiden_partition(adjacency) if method == "leiden" else self.louvain_partition(adjacency)
        groups: dict[str, set[str]] = defaultdict(set)
        for borrower_id, community_id in partition.items():
            groups[community_id].add(borrower_id)
        communities = [
            self.user_community_metrics(community_id, borrowers, adjacency, basis)
            for community_id, borrowers in groups.items()
        ]
        return sorted(
            communities,
            key=lambda item: (item["risk_score"], item["bad_debt_rate"], item["borrower_count"]),
            reverse=True,
        )[:limit]

    def agent_louvain_leiden_communities(
        self,
        method: str,
        basis: str | None = None,
        min_agents: int = 1,
        limit: int = 20,
    ) -> list[dict[str, Any]]:
        adjacency = self.agent_multi_projection()
        partition = self.leiden_partition(adjacency) if method == "leiden" else self.louvain_partition(adjacency)
        groups: dict[str, set[str]] = defaultdict(set)
        for agent_id, community_id in partition.items():
            groups[community_id].add(agent_id)
        communities = [
            self.community_metrics(community_id, agents, adjacency, basis)
            for community_id, agents in groups.items()
            if len(agents) >= min_agents
        ]
        return sorted(
            communities,
            key=lambda item: (item["risk_score"], item["bad_debt_rate"], item["agent_count"], item["borrower_count"]),
            reverse=True,
        )[:limit]

    def communities(
        self,
        method: str = "connected_components",
        basis: str | None = None,
        limit: int = 20,
    ) -> list[dict[str, Any]]:
        if method in {"louvain", "leiden"}:
            return self.agent_louvain_leiden_communities(method, basis, min_agents=5, limit=limit)
        return self.agent_communities(basis=basis, min_agents=5, limit=limit)

    def agent_communities(
        self,
        basis: str | None = None,
        min_shared_borrowers: int = 1,
        min_agents: int = 5,
        limit: int = 20,
    ) -> list[dict[str, Any]]:
        adjacency = self.filtered_agent_projection()
        components = self.agent_community_components(min_shared_borrowers, min_agents)
        communities = [
            self.community_metrics(f"C{index:03d}", agents, adjacency, basis)
            for index, agents in enumerate(components, 1)
        ]
        return sorted(
            communities,
            key=lambda item: (
                item["risk_score"],
                item["bad_debt_rate"],
                item["agent_count"],
                item["shared_borrower_count"],
            ),
            reverse=True,
        )[:limit]

    def community_feature_index(self, basis: str | None = None, method: str = "connected_components") -> dict[str, Any]:
        adjacency = self.agent_projection()
        connected_adjacency = self.filtered_agent_projection()
        multi_adjacency = self.agent_multi_projection()
        device_adjacency = self.agent_projection_by_column(DEVICE_COL)
        ip_adjacency = self.agent_projection_by_column(IP_COL)
        address_adjacency = self.agent_projection_by_column(ADDR_CLUSTER_COL)
        pagerank = self.pagerank(adjacency)
        connected_pagerank = self.pagerank(connected_adjacency)
        multi_pagerank = self.pagerank(multi_adjacency)
        agent_features: dict[str, dict[str, Any]] = {}
        community_by_agent: dict[str, dict[str, Any]] = {}
        community_by_borrower: dict[str, dict[str, Any]] = {}
        agent_community_by_agent: dict[str, dict[str, Any]] = {}
        agent_community_by_id: dict[str, dict[str, Any]] = {}

        for index, agents in enumerate(self.agent_community_components(), 1):
            community = self.community_metrics(f"C{index:03d}", agents, connected_adjacency, basis)
            betweenness = self.community_betweenness(agents, connected_adjacency)
            for agent_id in agents:
                degree = sum(1 for neighbor in connected_adjacency.get(agent_id, {}) if neighbor in agents)
                weighted_degree = sum(
                    weight
                    for neighbor, weight in connected_adjacency.get(agent_id, {}).items()
                    if neighbor in agents
                )
                agent_features[agent_id] = {
                    "节点度数(连通分量关系)": degree,
                    "加权度数(连通分量关系)": weighted_degree,
                    "PageRank分数(连通分量关系)": connected_pagerank.get(agent_id, 0),
                    "中介中心性(连通分量关系)": betweenness.get(agent_id, 0),
                    "连通子图大小(连通分量关系)": len(agents),
                    "社区id(连通分量关系)": community["community_id"],
                }
                community_by_agent[agent_id] = community

        for agent_id in self.agent_to_borrowers:
            if agent_id in agent_features:
                continue
            agent_loans = self.funded_loans(self.loans_for_borrowers(self.agent_to_borrowers.get(agent_id, set())))
            agent_features[agent_id] = {
                "节点度数(连通分量关系)": len(connected_adjacency.get(agent_id, {})),
                "加权度数(连通分量关系)": sum(connected_adjacency.get(agent_id, {}).values()),
                "PageRank分数(连通分量关系)": connected_pagerank.get(agent_id, 0),
                "中介中心性(连通分量关系)": 0,
                "连通子图大小(连通分量关系)": 1,
                "社区id(连通分量关系)": "",
            }
            community_by_agent[agent_id] = {
                "community_id": "",
                "agent_count": 1,
                "borrower_count": len(self.agent_to_borrowers.get(agent_id, set())),
                "loan_count": len(agent_loans),
                "bad_debt_rate": self.bad_rate(agent_loans, basis),
                "device_density": 0,
                "ip_density": 0,
                "address_density": 0,
                "risk_score": 0,
                "shared_borrower_count": 0,
            }

        shared_user_seen: set[str] = set()
        for index, agents in enumerate(self.agent_components_from_adjacency(adjacency, min_agents=2), 1):
            betweenness = self.community_betweenness(agents, adjacency)
            community_id = f"SC{index:03d}"
            for agent_id in agents:
                degree = sum(1 for neighbor in adjacency.get(agent_id, {}) if neighbor in agents)
                weighted_degree = sum(
                    weight
                    for neighbor, weight in adjacency.get(agent_id, {}).items()
                    if neighbor in agents
                )
                features = agent_features.setdefault(agent_id, {})
                features.update(
                    {
                        "节点度数(共享用户)": degree,
                        "加权度数(共享用户)": weighted_degree,
                        "PageRank分数(共享用户)": pagerank.get(agent_id, 0),
                        "中介中心性(共享用户)": betweenness.get(agent_id, 0),
                        "连通子图大小(共享用户)": len(agents),
                        "社区id(共享用户)": community_id,
                    }
                )
                shared_user_seen.add(agent_id)

        for agent_id in self.agent_to_borrowers:
            if agent_id in shared_user_seen:
                continue
            features = agent_features.setdefault(agent_id, {})
            features.update(
                {
                    "节点度数(共享用户)": len(adjacency.get(agent_id, {})),
                    "加权度数(共享用户)": sum(adjacency.get(agent_id, {}).values()),
                    "PageRank分数(共享用户)": pagerank.get(agent_id, 0),
                    "中介中心性(共享用户)": 0,
                    "连通子图大小(共享用户)": 1,
                    "社区id(共享用户)": "",
                }
            )

        relation_adjacencies = {
            "共享设备": device_adjacency,
            "共享IP": ip_adjacency,
            "共享地址簇": address_adjacency,
        }
        for agent_id in self.agent_to_borrowers:
            features = agent_features.setdefault(agent_id, {})
            for label, relation_adjacency in relation_adjacencies.items():
                features[f"节点度数({label})"] = len(relation_adjacency.get(agent_id, {}))
                features[f"加权度数({label})"] = sum(relation_adjacency.get(agent_id, {}).values())

        for index, agents in enumerate(self.agent_components_from_adjacency(multi_adjacency), 1):
            betweenness = self.community_betweenness(agents, multi_adjacency)
            for agent_id in agents:
                neighbors = {
                    neighbor: weight
                    for neighbor, weight in multi_adjacency.get(agent_id, {}).items()
                    if neighbor in agents
                }
                features = agent_features.setdefault(agent_id, {})
                features.update(
                    {
                        "节点度数(多关系)": len(neighbors),
                        "加权度数(多关系)": sum(neighbors.values()),
                        "PageRank分数(多关系)": multi_pagerank.get(agent_id, 0),
                        "中介中心性(多关系)": betweenness.get(agent_id, 0),
                        "连通子图大小(多关系)": len(agents),
                        "社区id(多关系)": f"MC{index:03d}" if len(agents) >= 2 else "",
                    }
                )

        if method in {"louvain", "leiden"}:
            user_adjacency = self.user_projection()
            partition = self.leiden_partition(user_adjacency) if method == "leiden" else self.louvain_partition(user_adjacency)
            groups: dict[str, set[str]] = defaultdict(set)
            for borrower_id, community_id in partition.items():
                groups[community_id].add(borrower_id)
            for community_id, borrowers in groups.items():
                community = self.user_community_metrics(community_id, borrowers, user_adjacency, basis)
                for borrower_id in borrowers:
                    community_by_borrower[borrower_id] = community

            agent_adjacency = self.agent_multi_projection()
            agent_partition = self.leiden_partition(agent_adjacency) if method == "leiden" else self.louvain_partition(agent_adjacency)
            agent_groups: dict[str, set[str]] = defaultdict(set)
            for agent_id, community_id in agent_partition.items():
                agent_groups[community_id].add(agent_id)
            agent_pagerank = self.pagerank(agent_adjacency)
            for community_id, agents in agent_groups.items():
                community = self.community_metrics(community_id, agents, agent_adjacency, basis)
                betweenness = self.community_betweenness(agents, agent_adjacency)
                for agent_id in agents:
                    degree = sum(1 for neighbor in agent_adjacency.get(agent_id, {}) if neighbor in agents)
                    weighted_degree = sum(
                        weight
                        for neighbor, weight in agent_adjacency.get(agent_id, {}).items()
                        if neighbor in agents
                    )
                    agent_features.setdefault(agent_id, {}).update(
                        {
                            "节点度数(Agent)": degree,
                            "加权度数(Agent)": weighted_degree,
                            "PageRank分数(Agent)": agent_pagerank.get(agent_id, 0),
                            "中介中心性(Agent)": betweenness.get(agent_id, 0),
                            "社区id(Agent)": community_id,
                            "社区规模(Agent)": community["agent_count"],
                            "社区借款人数(Agent)": community["borrower_count"],
                            "社区贷款笔数(Agent)": community["loan_count"],
                            "社区坏账率(Agent)": community["bad_debt_rate"],
                            "社区设备密度(Agent)": community["device_density"],
                            "社区ip密度(Agent)": community["ip_density"],
                            "社区地址密度(Agent)": community["address_density"],
                            "社区共享借款人数(Agent)": community["shared_borrower_count"],
                            "社区风险分(Agent)": community["risk_score"],
                        }
                    )
                    agent_community_by_agent[agent_id] = community
                agent_community_by_id[community_id] = community

        return {
            "adjacency": adjacency,
            "multi_adjacency": multi_adjacency,
            "agent_features": agent_features,
            "community_by_agent": community_by_agent,
            "community_by_borrower": community_by_borrower,
            "agent_community_by_agent": agent_community_by_agent,
            "agent_community_by_id": agent_community_by_id,
        }

    def feature_table(self, basis: str | None = None, community_method: str = "connected_components") -> dict[str, Any]:
        loans_by_agent: dict[str, list[dict[str, Any]]] = defaultdict(list)
        loans_by_addr_cluster: dict[str, list[dict[str, Any]]] = defaultdict(list)
        loans_by_device: dict[str, list[dict[str, Any]]] = defaultdict(list)
        loans_by_ip: dict[str, list[dict[str, Any]]] = defaultdict(list)

        for loan in self.loan_rows:
            agent_id = normalize_phone(loan.get(AGENT_COL))
            addr_cluster_id = normalize_value(loan.get(ADDR_CLUSTER_COL))
            device_id = normalize_value(loan.get(DEVICE_COL))
            ip = normalize_value(loan.get(IP_COL))
            if agent_id:
                loans_by_agent[agent_id].append(loan)
            if addr_cluster_id:
                loans_by_addr_cluster[addr_cluster_id].append(loan)
            if device_id:
                loans_by_device[device_id].append(loan)
            if ip:
                loans_by_ip[ip].append(loan)

        agent_stats = {key: self.loan_group_stats(value, basis) for key, value in loans_by_agent.items()}
        addr_stats = {key: self.loan_group_stats(value, basis) for key, value in loans_by_addr_cluster.items()}
        device_stats = {key: self.loan_group_stats(value, basis) for key, value in loans_by_device.items()}
        ip_stats = {key: self.loan_group_stats(value, basis) for key, value in loans_by_ip.items()}
        borrower_funded_loans_by_id = {
            borrower_id: self.funded_loans(loans)
            for borrower_id, loans in self.loans_by_borrower.items()
        }
        borrower_bad_rate_by_id = {
            borrower_id: self.bad_rate(funded_loans, basis)
            for borrower_id, funded_loans in borrower_funded_loans_by_id.items()
        }
        second_degree_cache: dict[str, dict[str, Any]] = {}

        def second_degree_stats(borrower_id: str) -> dict[str, Any]:
            if borrower_id in second_degree_cache:
                return second_degree_cache[borrower_id]
            borrowers = self.second_degree_borrowers(borrower_id)
            loans = self.loans_for_borrowers(borrowers)
            funded = self.funded_loans(loans)
            stats = {
                "borrowers": borrowers,
                "loans": loans,
                "funded_loans": funded,
                "devices": self.distinct_from_loans(loans, DEVICE_COL),
                "ips": self.distinct_from_loans(loans, IP_COL),
                "addr_clusters": self.distinct_from_loans(loans, ADDR_CLUSTER_COL),
                "application_borrowers": {loan_item[BORROWER_COL] for loan_item in loans if loan_item.get(BORROWER_COL)},
                "bad_rate": self.bad_rate(funded, basis),
            }
            second_degree_cache[borrower_id] = stats
            return stats

        graph_features = self.community_feature_index(basis, community_method)
        agent_features = graph_features["agent_features"]
        community_by_agent = graph_features["community_by_agent"]
        community_by_borrower = graph_features["community_by_borrower"]
        agent_community_by_agent = graph_features["agent_community_by_agent"]
        selected_community_suffix = "User" if community_method in {"louvain", "leiden"} else "Agent"
        selected_community_columns = [
            f"社区id({selected_community_suffix})",
            f"社区规模({selected_community_suffix})",
            f"社区借款人数({selected_community_suffix})",
            f"社区贷款笔数({selected_community_suffix})",
            f"社区坏账率({selected_community_suffix})",
            f"社区设备密度({selected_community_suffix})",
            f"社区ip密度({selected_community_suffix})",
            f"社区地址密度({selected_community_suffix})",
            f"社区共享借款人数({selected_community_suffix})",
            f"社区风险分({selected_community_suffix})",
        ]

        feature_columns = [
            "地址簇大小_地址数",
            "收货人手机号关联人数",
            "收货人手机号关联设备数",
            "收货人手机号关联ip数",
            "收货人手机号关联地址簇数",
            "收货人手机号坏账率",
            "地址簇关联人数",
            "地址簇关联收货人手机号数",
            "地址簇关联设备数",
            "地址簇关联ip数",
            "地址簇申请笔数",
            "地址簇坏账率",
            "设备关联人数",
            "设备关联地址簇数",
            "设备关联收货手机号数",
            "设备申请笔数",
            "设备坏账率",
            "IP关联人数",
            "IP关联设备数",
            "IP申请笔数",
            "IP坏账率",
            "借款人一度中介数",
            "借款人贷款笔数",
            "借款人坏账率",
            "二度借款人数",
            "二度中介数",
            "二度设备数",
            "二度ip数",
            "二度地址簇数",
            "二度坏账率",
            "二度申请人数",
            "二度申请笔数",
            "二度贷款笔数",
            "节点度数(共享用户)",
            "加权度数(共享用户)",
            "PageRank分数(共享用户)",
            "中介中心性(共享用户)",
            "连通子图大小(共享用户)",
            "社区id(共享用户)",
            "节点度数(连通分量关系)",
            "加权度数(连通分量关系)",
            "PageRank分数(连通分量关系)",
            "中介中心性(连通分量关系)",
            "连通子图大小(连通分量关系)",
            "社区id(连通分量关系)",
            "节点度数(共享设备)",
            "加权度数(共享设备)",
            "节点度数(共享IP)",
            "加权度数(共享IP)",
            "节点度数(共享地址簇)",
            "加权度数(共享地址簇)",
            "节点度数(多关系)",
            "加权度数(多关系)",
            "PageRank分数(多关系)",
            "中介中心性(多关系)",
            "连通子图大小(多关系)",
            "社区id(多关系)",
            *selected_community_columns,
        ]
        if community_method in {"louvain", "leiden"}:
            feature_columns += [
                "节点度数(Agent)",
                "加权度数(Agent)",
                "PageRank分数(Agent)",
                "中介中心性(Agent)",
                "社区id(Agent)",
                "社区规模(Agent)",
                "社区借款人数(Agent)",
                "社区贷款笔数(Agent)",
                "社区坏账率(Agent)",
                "社区设备密度(Agent)",
                "社区ip密度(Agent)",
                "社区地址密度(Agent)",
                "社区共享借款人数(Agent)",
                "社区风险分(Agent)",
            ]

        rows: list[dict[str, Any]] = []
        for loan in self.loan_rows:
            agent_id = normalize_phone(loan.get(AGENT_COL))
            borrower_id = normalize_phone(loan.get(BORROWER_COL))
            addr_cluster_id = normalize_value(loan.get(ADDR_CLUSTER_COL))
            device_id = normalize_value(loan.get(DEVICE_COL))
            ip = normalize_value(loan.get(IP_COL))

            agent_stat = agent_stats.get(agent_id, {})
            addr_stat = addr_stats.get(addr_cluster_id, {})
            device_stat = device_stats.get(device_id, {})
            ip_stat = ip_stats.get(ip, {})
            second_stat = second_degree_stats(borrower_id)
            second_borrowers = second_stat["borrowers"]
            second_agents = self.second_degree_agents(agent_id)
            second_loans = second_stat["loans"]
            second_funded_loans = second_stat["funded_loans"]
            borrower_funded_loans = borrower_funded_loans_by_id.get(borrower_id, [])
            community = (
                community_by_borrower.get(borrower_id, {})
                if community_method in {"louvain", "leiden"}
                else community_by_agent.get(agent_id, {})
            )
            network = agent_features.get(agent_id, {})
            selected_community_id = (
                community.get("community_id", "")
                if community_method in {"louvain", "leiden"}
                else network.get("社区id(连通分量关系)", "")
            )
            selected_community_size = (
                community.get("borrower_count", 0)
                if community_method in {"louvain", "leiden"}
                else community.get("agent_count", 1)
            )

            feature_values = {
                "地址簇大小_地址数": len(addr_stat.get("addresses", set())),
                "收货人手机号关联人数": len(agent_stat.get("borrowers", set())),
                "收货人手机号关联设备数": len(agent_stat.get("devices", set())),
                "收货人手机号关联ip数": len(agent_stat.get("ips", set())),
                "收货人手机号关联地址簇数": len(agent_stat.get("addr_clusters", set())),
                "收货人手机号坏账率": agent_stat.get("bad_rate", 0),
                "地址簇关联人数": len(addr_stat.get("borrowers", set())),
                "地址簇关联收货人手机号数": len(addr_stat.get("agents", set())),
                "地址簇关联设备数": len(addr_stat.get("devices", set())),
                "地址簇关联ip数": len(addr_stat.get("ips", set())),
                "地址簇申请笔数": addr_stat.get("application_count", 0),
                "地址簇坏账率": addr_stat.get("bad_rate", 0),
                "设备关联人数": len(device_stat.get("borrowers", set())),
                "设备关联地址簇数": len(device_stat.get("addr_clusters", set())),
                "设备关联收货手机号数": len(device_stat.get("agents", set())),
                "设备申请笔数": device_stat.get("application_count", 0),
                "设备坏账率": device_stat.get("bad_rate", 0),
                "IP关联人数": len(ip_stat.get("borrowers", set())),
                "IP关联设备数": len(ip_stat.get("devices", set())),
                "IP申请笔数": ip_stat.get("application_count", 0),
                "IP坏账率": ip_stat.get("bad_rate", 0),
                "借款人一度中介数": len(self.first_degree_agents(borrower_id)),
                "借款人贷款笔数": len(borrower_funded_loans),
                "借款人坏账率": borrower_bad_rate_by_id.get(borrower_id, 0),
                "二度借款人数": len(second_borrowers),
                "二度中介数": len(second_agents),
                "二度设备数": len(second_stat["devices"]),
                "二度ip数": len(second_stat["ips"]),
                "二度地址簇数": len(second_stat["addr_clusters"]),
                "二度坏账率": second_stat["bad_rate"],
                "二度申请人数": len(second_stat["application_borrowers"]),
                "二度申请笔数": len(second_loans),
                "二度贷款笔数": len(second_funded_loans),
                "节点度数(共享用户)": network.get("节点度数(共享用户)", 0),
                "加权度数(共享用户)": network.get("加权度数(共享用户)", 0),
                "PageRank分数(共享用户)": network.get("PageRank分数(共享用户)", 0),
                "中介中心性(共享用户)": network.get("中介中心性(共享用户)", 0),
                "连通子图大小(共享用户)": network.get("连通子图大小(共享用户)", 1),
                "社区id(共享用户)": network.get("社区id(共享用户)", ""),
                "节点度数(连通分量关系)": network.get("节点度数(连通分量关系)", 0),
                "加权度数(连通分量关系)": network.get("加权度数(连通分量关系)", 0),
                "PageRank分数(连通分量关系)": network.get("PageRank分数(连通分量关系)", 0),
                "中介中心性(连通分量关系)": network.get("中介中心性(连通分量关系)", 0),
                "连通子图大小(连通分量关系)": network.get("连通子图大小(连通分量关系)", 1),
                "社区id(连通分量关系)": network.get("社区id(连通分量关系)", ""),
                "节点度数(共享设备)": network.get("节点度数(共享设备)", 0),
                "加权度数(共享设备)": network.get("加权度数(共享设备)", 0),
                "节点度数(共享IP)": network.get("节点度数(共享IP)", 0),
                "加权度数(共享IP)": network.get("加权度数(共享IP)", 0),
                "节点度数(共享地址簇)": network.get("节点度数(共享地址簇)", 0),
                "加权度数(共享地址簇)": network.get("加权度数(共享地址簇)", 0),
                "节点度数(多关系)": network.get("节点度数(多关系)", 0),
                "加权度数(多关系)": network.get("加权度数(多关系)", 0),
                "PageRank分数(多关系)": network.get("PageRank分数(多关系)", 0),
                "中介中心性(多关系)": network.get("中介中心性(多关系)", 0),
                "连通子图大小(多关系)": network.get("连通子图大小(多关系)", 1),
                "社区id(多关系)": network.get("社区id(多关系)", ""),
                f"社区id({selected_community_suffix})": selected_community_id,
                f"社区规模({selected_community_suffix})": selected_community_size,
                f"社区借款人数({selected_community_suffix})": community.get("borrower_count", 0),
                f"社区贷款笔数({selected_community_suffix})": community.get("loan_count", 0),
                f"社区坏账率({selected_community_suffix})": community.get("bad_debt_rate", 0),
                f"社区设备密度({selected_community_suffix})": community.get("device_density", 0),
                f"社区ip密度({selected_community_suffix})": community.get("ip_density", 0),
                f"社区地址密度({selected_community_suffix})": community.get("address_density", 0),
                f"社区共享借款人数({selected_community_suffix})": community.get("shared_borrower_count", 0),
                f"社区风险分({selected_community_suffix})": community.get("risk_score", 0),
            }
            if community_method in {"louvain", "leiden"}:
                feature_values.update(
                    {
                        "节点度数(Agent)": network.get("节点度数(Agent)", 0),
                        "加权度数(Agent)": network.get("加权度数(Agent)", 0),
                        "PageRank分数(Agent)": network.get("PageRank分数(Agent)", 0),
                        "中介中心性(Agent)": network.get("中介中心性(Agent)", 0),
                        "社区id(Agent)": network.get("社区id(Agent)", ""),
                        "社区规模(Agent)": network.get("社区规模(Agent)", 0),
                        "社区借款人数(Agent)": network.get("社区借款人数(Agent)", 0),
                        "社区贷款笔数(Agent)": network.get("社区贷款笔数(Agent)", 0),
                        "社区坏账率(Agent)": network.get("社区坏账率(Agent)", 0),
                        "社区设备密度(Agent)": network.get("社区设备密度(Agent)", 0),
                        "社区ip密度(Agent)": network.get("社区ip密度(Agent)", 0),
                        "社区地址密度(Agent)": network.get("社区地址密度(Agent)", 0),
                        "社区共享借款人数(Agent)": network.get("社区共享借款人数(Agent)", 0),
                        "社区风险分(Agent)": network.get("社区风险分(Agent)", 0),
                    }
                )
            rows.append({**loan, **feature_values})

        columns = [column for column in self.source_columns if column in rows[0]] if rows else self.source_columns[:]
        for column in feature_columns:
            if column not in columns:
                columns.append(column)
        return {"columns": columns, "rows": rows}

    def summary(self, basis: str | None = None, community_method: str = "connected_components") -> dict[str, Any]:
        missing = sorted(REQUIRED_COLUMNS - set(self.columns))
        funded = self.funded_loans(self.loan_rows)
        overdue_total = sum(1 for loan in funded if self.is_overdue(loan, basis))
        if community_method in {"louvain", "leiden"}:
            communities = self.agent_louvain_leiden_communities(community_method, basis, limit=10_000)
            community_count = len(communities)
            max_community_size = max((item["agent_count"] for item in communities), default=0)
            max_community_agents = max_community_size
            community_size_label = "最大团伙中介数"
        else:
            components = self.agent_community_components()
            community_count = len(components)
            max_community_size = max((len(item) for item in components), default=0)
            max_community_agents = max_community_size
            community_size_label = "最大团伙中介数"
        return {
            "row_count": len(self.rows),
            "valid_loan_count": len(funded),
            "skipped_rows": self.skipped_rows,
            "agent_count": len(self.agent_to_borrowers),
            "borrower_count": len(self.borrower_to_agents),
            "relation_count": sum(len(v) for v in self.agent_to_borrowers.values()),
            "community_count": community_count,
            "max_community_agents": max_community_agents,
            "max_community_size": max_community_size,
            "community_size_label": community_size_label,
            "community_method": community_method,
            "overdue_loan_count": overdue_total,
            "overdue_rate": overdue_total / len(funded) if funded else 0,
            "missing_columns": missing,
            "columns": self.columns,
            **self.graph_quality_metrics(community_method),
        }

    def search(self, query: str, limit: int = 10) -> dict[str, list[dict[str, str]]]:
        needle = normalize_phone(query)
        if not needle:
            return {"agents": [], "borrowers": []}

        def matched(values: list[str], node_type: str) -> list[dict[str, str]]:
            hits = [value for value in values if needle in value]
            return [{"id": value, "label": mask_phone(value), "type": node_type} for value in hits[:limit]]

        return {
            "agents": matched(sorted(self.agent_to_borrowers), "agent"),
            "borrowers": matched(sorted(self.borrower_to_agents), "borrower"),
        }

    def graph(self, center_type: str, center_id: str, basis: str | None = None, second_limit: int = 160) -> dict[str, Any]:
        center_id = normalize_phone(center_id)
        center_type = "agent" if center_type == "agent" else "borrower"
        nodes: dict[str, dict[str, Any]] = {}
        edges: set[tuple[str, str]] = set()

        def add_node(raw_id: str, node_type: str, level: int) -> None:
            if not raw_id:
                return
            node_id = f"{node_type}:{raw_id}"
            if node_id not in nodes:
                label = mask_phone(raw_id) if node_type in {"agent", "borrower"} else raw_id
                data: dict[str, Any] = {
                    "id": node_id,
                    "raw_id": raw_id,
                    "type": node_type,
                    "label": label,
                    "level": level,
                }
                if node_type == "borrower":
                    own_loans = self.funded_loans(self.loans_by_borrower.get(raw_id, []))
                    data["loan_count"] = len(own_loans)
                    data["overdue"] = any(self.is_overdue(loan, basis) for loan in own_loans)
                nodes[node_id] = data
            else:
                nodes[node_id]["level"] = min(nodes[node_id]["level"], level)

        def add_edge(agent_id: str, borrower_id: str) -> None:
            edges.add((f"agent:{agent_id}", f"borrower:{borrower_id}", "agent_borrower"))

        def add_raw_edge(source_type: str, source_id: str, target_type: str, target_id: str) -> None:
            if source_id and target_id:
                edges.add((f"{source_type}:{source_id}", f"{target_type}:{target_id}", f"{source_type}_{target_type}"))

        def add_context_nodes(agent_id: str, borrower_id: str, level: int) -> None:
            linked_loans = [
                loan
                for loan in self.loans_by_borrower.get(borrower_id, [])
                if normalize_phone(loan.get(AGENT_COL)) == agent_id
            ]
            for loan in linked_loans:
                device_id = normalize_value(loan.get(DEVICE_COL))
                ip = normalize_value(loan.get(IP_COL))
                addr_cluster_id = normalize_value(loan.get(ADDR_CLUSTER_COL))
                if device_id:
                    add_node(device_id, "device", level)
                    add_raw_edge("borrower", borrower_id, "device", device_id)
                if ip:
                    add_node(ip, "ip", level)
                    add_raw_edge("borrower", borrower_id, "ip", ip)
                if addr_cluster_id:
                    add_node(addr_cluster_id, "address", level)
                    add_raw_edge("agent", agent_id, "address", addr_cluster_id)

        if center_type == "agent":
            add_node(center_id, "agent", 0)
            first = sorted(self.first_degree_borrowers(center_id))
            second = sorted(
                self.second_degree_agents(center_id),
                key=lambda item: len(self.first_degree_borrowers(item)),
                reverse=True,
            )[:second_limit]
            second_set = set(second)
            for borrower_id in first:
                add_node(borrower_id, "borrower", 1)
                add_edge(center_id, borrower_id)
                add_context_nodes(center_id, borrower_id, 2)
                for agent_id in sorted(self.borrower_to_agents.get(borrower_id, set())):
                    if agent_id in second_set:
                        add_node(agent_id, "agent", 2)
                        add_edge(agent_id, borrower_id)
                        add_context_nodes(agent_id, borrower_id, 3)
            metrics = self.agent_metrics(center_id)
        else:
            add_node(center_id, "borrower", 0)
            first = sorted(self.first_degree_agents(center_id))
            second = sorted(
                self.second_degree_borrowers(center_id),
                key=lambda item: self.second_degree_overdue_rate(item, basis)["second_degree_overdue_rate"],
                reverse=True,
            )[:second_limit]
            second_set = set(second)
            for agent_id in first:
                add_node(agent_id, "agent", 1)
                add_edge(agent_id, center_id)
                add_context_nodes(agent_id, center_id, 2)
                for borrower_id in sorted(self.agent_to_borrowers.get(agent_id, set())):
                    if borrower_id in second_set:
                        add_node(borrower_id, "borrower", 2)
                        add_edge(agent_id, borrower_id)
                        add_context_nodes(agent_id, borrower_id, 3)
            metrics = self.borrower_metrics(center_id, basis)

        return {
            "center": {"type": center_type, "id": center_id, "node_id": f"{center_type}:{center_id}"},
            "metrics": metrics,
            "nodes": list(nodes.values()),
            "edges": [
                {"source": source, "target": target, "relation": relation}
                for source, target, relation in sorted(edges)
            ],
            "truncated": len(nodes) >= second_limit,
        }

    def default_center(self, basis: str | None = None) -> tuple[str, str]:
        borrowers = self.top_borrowers(basis, 1)
        if borrowers:
            return "borrower", borrowers[0]["borrower_id"]
        agents = self.top_agents(1)
        if agents:
            return "agent", agents[0]["agent_id"]
        return "borrower", ""


class Handler(BaseHTTPRequestHandler):
    def log_message(self, fmt: str, *args: Any) -> None:
        print(f"{self.address_string()} - {fmt % args}")

    def do_GET(self) -> None:  # noqa: N802 - stdlib API
        path = self.path.split("?", 1)[0]
        if path == "/":
            path = "/index.html"
        file_path = (STATIC / path.lstrip("/")).resolve()
        if not str(file_path).startswith(str(STATIC.resolve())) or not file_path.exists():
            self.send_error(404)
            return
        content = file_path.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", mimetypes.guess_type(file_path.name)[0] or "application/octet-stream")
        self.send_header("Content-Length", str(len(content)))
        self.end_headers()
        self.wfile.write(content)

    def do_POST(self) -> None:  # noqa: N802 - stdlib API
        length = int(self.headers.get("Content-Length", "0"))
        try:
            payload = json.loads(self.rfile.read(length).decode("utf-8"))
            if self.path == "/api/analyze":
                self.handle_analyze(payload)
            elif self.path == "/api/graph":
                self.handle_graph(payload)
            elif self.path == "/api/search":
                self.handle_search(payload)
            elif self.path == "/api/complex-query":
                self.handle_complex_query(payload)
            elif self.path == "/api/cluster/full":
                self.handle_cluster_full(payload)
            elif self.path == "/api/cluster/incremental":
                self.handle_cluster_incremental(payload)
            elif self.path == "/api/cluster/status":
                self.handle_cluster_status(payload)
            elif self.path == "/api/cluster/result":
                self.handle_cluster_result(payload)
            else:
                json_response(self, 404, {"error": "接口不存在"})
        except Exception as exc:  # noqa: BLE001 - convert to JSON for the UI
            json_response(self, 400, {"error": str(exc)})

    def handle_analyze(self, payload: dict[str, Any]) -> None:
        filename = payload.get("filename") or ""
        basis = payload.get("overdue_basis") or "any"
        community_method = payload.get("community_method") or "connected_components"
        original_rows = parse_uploaded_file(filename, payload.get("content_base64") or "")
        rows = filter_rows(original_rows, payload.get("reloan_filter") or "all", payload.get("return_filter") or "all")
        reloan_filter = payload.get("reloan_filter") or "all"
        graph = RiskGraph(rows, basis, reloan_filter)
        dataset_id = uuid.uuid4().hex
        DATASETS[dataset_id] = {"graph": graph, "original_rows": original_rows}
        center_type, center_id = graph.default_center(basis)
        feature_table = graph.feature_table(basis, community_method)
        summary = graph.summary(basis, community_method)
        summary.update(summary_from_feature_table(feature_table))
        summary["uploaded_row_count"] = len(original_rows)
        summary["analyzed_row_count"] = len(rows)
        json_response(
            self,
            200,
            {
                "dataset_id": dataset_id,
                "summary": summary,
                "filter_options": {
                    **build_filter_options(original_rows),
                },
                "top_agents": graph.top_agents(),
                "top_borrowers": graph.top_borrowers(basis),
                "communities": graph.communities(community_method, basis),
                "feature_table": feature_table,
                "graph": graph.graph(center_type, center_id, basis) if center_id else {"nodes": [], "edges": [], "metrics": {}},
            },
        )

    def dataset_entry(self, payload: dict[str, Any]) -> dict[str, Any]:
        dataset_id = payload.get("dataset_id") or ""
        if dataset_id not in DATASETS:
            raise ValueError("数据集不存在，请重新上传文件。")
        return DATASETS[dataset_id]

    def dataset(self, payload: dict[str, Any]) -> RiskGraph:
        return self.dataset_entry(payload)["graph"]

    def handle_graph(self, payload: dict[str, Any]) -> None:
        graph = self.dataset(payload)
        basis = payload.get("overdue_basis") or graph.overdue_basis
        center_type = payload.get("center_type") or "borrower"
        center_id = payload.get("center_id") or ""
        json_response(self, 200, {"graph": graph.graph(center_type, center_id, basis)})

    def handle_search(self, payload: dict[str, Any]) -> None:
        graph = self.dataset(payload)
        json_response(self, 200, graph.search(payload.get("query") or ""))

    def handle_complex_query(self, payload: dict[str, Any]) -> None:
        entry = self.dataset_entry(payload)
        result = complex_query_rows(
            entry["original_rows"],
            payload.get("filters") or {},
            payload.get("query_values") or {},
        )
        json_response(self, 200, result)

    def handle_cluster_full(self, payload: dict[str, Any]) -> None:
        filename = payload.get("filename") or ""
        rows = parse_uploaded_file(filename, payload.get("content_base64") or "")
        params = payload.get("params") or {}
        job_id = start_cluster_job("full", run_full_cluster_job, rows, params)
        json_response(self, 200, {"job_id": job_id})

    def handle_cluster_incremental(self, payload: dict[str, Any]) -> None:
        incremental_filename = payload.get("incremental_filename") or ""
        incremental_rows = parse_uploaded_file(incremental_filename, payload.get("incremental_content_base64") or "")
        base_content = base64.b64decode(payload.get("base_content_base64") or "")
        try:
            base_payload = json.loads(base_content.decode("utf-8"))
        except Exception as exc:  # noqa: BLE001 - invalid uploaded JSON
            raise ValueError(f"工程文件解析失败：{exc}") from exc
        params = payload.get("params") or {}
        job_id = start_cluster_job("incremental", run_incremental_cluster_job, base_payload, incremental_rows, params)
        json_response(self, 200, {"job_id": job_id})

    def handle_cluster_status(self, payload: dict[str, Any]) -> None:
        json_response(self, 200, cluster_job(payload.get("job_id") or ""))

    def handle_cluster_result(self, payload: dict[str, Any]) -> None:
        json_response(self, 200, cluster_result(payload.get("job_id") or ""))


def main() -> None:
    print(f"ANTI-FRAUD graph server running at http://{HOST}:{PORT}")
    ThreadingHTTPServer((HOST, PORT), Handler).serve_forever()


if __name__ == "__main__":
    main()
